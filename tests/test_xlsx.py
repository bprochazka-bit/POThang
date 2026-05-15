"""Tests for the xlsx PO template renderer."""
from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook, load_workbook

from purchasetracker.models import Item, POLine, PurchaseOrder
from purchasetracker.services import render_po_xlsx


def _build_template_bytes() -> bytes:
    """Build a minimal in-memory template with both named cells and a loop."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "PO: {{po_number}}"
    ws["A2"] = "Vendor: {{vendor}}"
    ws["A3"] = "Date: {{date}}"
    # Loop region
    ws["A5"] = "{{#items}}"
    ws["A6"] = "{{item.index}}"
    ws["B6"] = "{{item.description}}"
    ws["C6"] = "{{item.qty}}"
    ws["D6"] = "{{item.unit_cost}}"
    ws["E6"] = "{{item.line_total}}"
    ws["A7"] = "{{/items}}"
    ws["A9"] = "Total: {{total}}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _seed_po(db) -> PurchaseOrder:
    item_a = Item(description="Widget A", qty=2, unit_cost=10.5)
    item_b = Item(description="Widget B", qty=3, unit_cost=2.0)
    db.session.add_all([item_a, item_b])
    db.session.flush()
    po = PurchaseOrder(po_number="PO-2026-001", vendor="Acme Co.")
    db.session.add(po)
    db.session.flush()
    db.session.add(POLine(po_id=po.id, item_id=item_a.id, qty=2, unit_cost=10.5))
    db.session.add(POLine(po_id=po.id, item_id=item_b.id, qty=3, unit_cost=2.0))
    db.session.commit()
    return po


def test_named_cells_substituted(app, db):
    po = _seed_po(db)
    rendered = render_po_xlsx(po, _build_template_bytes())
    wb = load_workbook(io.BytesIO(rendered))
    ws = wb.active
    assert ws["A1"].value == "PO: PO-2026-001"
    assert ws["A2"].value == "Vendor: Acme Co."


def test_loop_expands_to_one_row_per_line(app, db):
    po = _seed_po(db)
    rendered = render_po_xlsx(po, _build_template_bytes())
    wb = load_workbook(io.BytesIO(rendered))
    ws = wb.active

    # Open marker at row 5, single template row at 6, close marker at 7.
    # After rendering: rows 5 and 6 should be the two line items, with
    # everything below shifted appropriately.
    assert ws.cell(row=5, column=1).value == 1
    assert ws.cell(row=5, column=2).value == "Widget A"
    assert ws.cell(row=5, column=3).value == 2
    assert ws.cell(row=5, column=4).value == 10.5
    assert ws.cell(row=5, column=5).value == 21.0

    assert ws.cell(row=6, column=1).value == 2
    assert ws.cell(row=6, column=2).value == "Widget B"
    assert ws.cell(row=6, column=3).value == 3
    assert ws.cell(row=6, column=4).value == 2.0
    assert ws.cell(row=6, column=5).value == 6.0


def test_total_named_cell_substituted(app, db):
    po = _seed_po(db)
    rendered = render_po_xlsx(po, _build_template_bytes())
    wb = load_workbook(io.BytesIO(rendered))
    ws = wb.active
    # The Total row started at A9 in the template; the loop expanded by 0 net
    # rows (deleted 3 marker+template rows, inserted 2 item rows = net -1),
    # so the total is at A8 now. Easier to find it by content scan.
    found = None
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "Total:" in c.value:
                found = c.value
                break
    assert found is not None
    assert "27" in found  # 21.0 + 6.0 = 27.0


def test_render_with_no_lines_drops_loop(app, db):
    po = PurchaseOrder(po_number="PO-empty", vendor="None")
    db.session.add(po)
    db.session.commit()
    rendered = render_po_xlsx(po, _build_template_bytes())
    wb = load_workbook(io.BytesIO(rendered))
    ws = wb.active
    # Markers and template row should all be gone, no blank line items remain.
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str):
                assert "{{#items}}" not in c.value
                assert "{{/items}}" not in c.value
                assert "{{item." not in c.value


def _build_conditional_template_bytes() -> bytes:
    """Template with an {{#if}} conditional in the loop row."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "{{#items}}"
    # Row 2 is the template row — between the markers.
    # Show model/sku when sku is present, otherwise just model.
    ws["B2"] = "{{#if item.vendor_sku}}{{item.model}}/{{item.vendor_sku}}{{else}}{{item.model}}{{/if}}"
    ws["A3"] = "{{/items}}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_conditional_both_fields_present(app, db):
    item = Item(name="Widget", model="MDL-1", vendor_sku="SKU-99")
    db.session.add(item)
    db.session.flush()
    po = PurchaseOrder(po_number="PO-COND-1", vendor="Acme")
    db.session.add(po)
    db.session.flush()
    db.session.add(POLine(po_id=po.id, item_id=item.id, qty=1, unit_cost=5.0))
    db.session.commit()

    rendered = render_po_xlsx(po, _build_conditional_template_bytes())
    ws = load_workbook(io.BytesIO(rendered)).active
    assert ws.cell(row=1, column=2).value == "MDL-1/SKU-99"


def test_conditional_sku_absent(app, db):
    item = Item(name="Widget", model="MDL-1")
    db.session.add(item)
    db.session.flush()
    po = PurchaseOrder(po_number="PO-COND-2", vendor="Acme")
    db.session.add(po)
    db.session.flush()
    db.session.add(POLine(po_id=po.id, item_id=item.id, qty=1, unit_cost=5.0))
    db.session.commit()

    rendered = render_po_xlsx(po, _build_conditional_template_bytes())
    ws = load_workbook(io.BytesIO(rendered)).active
    assert ws.cell(row=1, column=2).value == "MDL-1"


def test_conditional_if_only_no_else(app, db):
    """{{#if var}}text{{/if}} renders text when truthy, empty string when falsy."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "{{#items}}"
    ws["B2"] = "{{#if item.notes}}Note: {{item.notes}}{{/if}}"
    ws["A3"] = "{{/items}}"
    buf = io.BytesIO()
    wb.save(buf)
    tmpl = buf.getvalue()

    item_with_note = Item(name="A", notes=None)
    item_no_note = Item(name="B", notes=None)
    db.session.add_all([item_with_note, item_no_note])
    db.session.flush()
    po = PurchaseOrder(po_number="PO-COND-3", vendor="Acme")
    db.session.add(po)
    db.session.flush()
    line_with = POLine(po_id=po.id, item_id=item_with_note.id, qty=1, unit_cost=1.0, notes="fragile")
    line_without = POLine(po_id=po.id, item_id=item_no_note.id, qty=1, unit_cost=1.0)
    db.session.add_all([line_with, line_without])
    db.session.commit()

    rendered = render_po_xlsx(po, tmpl)
    ws = load_workbook(io.BytesIO(rendered)).active
    values = [ws.cell(row=r, column=2).value for r in (1, 2)]
    assert "Note: fragile" in values
    assert "" in values or None in values


def _merged_ranges(ws) -> set[str]:
    return {str(mr) for mr in ws.merged_cells.ranges}


def test_merged_cells_in_loop_replicated_per_line(app, db):
    """Merged cells inside the loop template are replicated for every rendered line."""
    wb = Workbook()
    ws = wb.active
    # Header row with a merge that must survive untouched.
    ws["A1"] = "Purchase Order"
    ws.merge_cells("A1:D1")
    # Loop region: marker, one template row with a merge, close marker.
    ws["A2"] = "{{#items}}"
    ws["A3"] = "{{item.name}}"
    ws.merge_cells("A3:B3")  # merge spans cols A-B in the template row
    ws["A4"] = "{{/items}}"
    # Footer merge that must shift down correctly.
    ws["A5"] = "Total"
    ws.merge_cells("A5:D5")
    buf = io.BytesIO()
    wb.save(buf)
    tmpl = buf.getvalue()

    item1 = Item(name="Alpha")
    item2 = Item(name="Beta")
    db.session.add_all([item1, item2])
    db.session.flush()
    po = PurchaseOrder(po_number="PO-MERGE-1", vendor="Acme")
    db.session.add(po)
    db.session.flush()
    db.session.add(POLine(po_id=po.id, item_id=item1.id, qty=1, unit_cost=1.0))
    db.session.add(POLine(po_id=po.id, item_id=item2.id, qty=1, unit_cost=1.0))
    db.session.commit()

    rendered = render_po_xlsx(po, tmpl)
    ws_out = load_workbook(io.BytesIO(rendered)).active
    ranges = _merged_ranges(ws_out)

    # Header merge must survive.
    assert "A1:D1" in ranges
    # Two lines → two A:B merges on rows 2 and 3.
    assert "A2:B2" in ranges
    assert "A3:B3" in ranges
    # Footer was at row 5 in the template; loop expanded by net 0 rows
    # (deleted 3 rows, inserted 2) so footer lands at row 4.
    assert "A4:D4" in ranges


def test_merged_cells_above_loop_unchanged(app, db):
    """Merged cells entirely above the loop block are not disturbed."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Header"
    ws.merge_cells("A1:C1")
    ws["A2"] = "{{#items}}"
    ws["A3"] = "{{item.name}}"
    ws["A4"] = "{{/items}}"
    buf = io.BytesIO()
    wb.save(buf)
    tmpl = buf.getvalue()

    item = Item(name="Widget")
    db.session.add(item)
    db.session.flush()
    po = PurchaseOrder(po_number="PO-MERGE-2", vendor="Acme")
    db.session.add(po)
    db.session.flush()
    db.session.add(POLine(po_id=po.id, item_id=item.id, qty=1, unit_cost=5.0))
    db.session.commit()

    rendered = render_po_xlsx(po, tmpl)
    ws_out = load_workbook(io.BytesIO(rendered)).active
    assert "A1:C1" in _merged_ranges(ws_out)


def _build_formula_template_bytes() -> bytes:
    """Template that uses Excel formulas with {{row}} and {{items.range.X}}."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "PO: {{po_number}}"
    ws["A3"] = "{{#items}}"
    # Template row 4: per-line formula referencing the current row.
    ws["A4"] = "{{item.description}}"
    ws["B4"] = "{{item.qty}}"
    ws["C4"] = "{{item.unit_cost}}"
    ws["D4"] = "=B{{row}}*C{{row}}"
    ws["A5"] = "{{/items}}"
    # Below the loop: aggregate over the expanded items range, plus tax.
    ws["A7"] = "Subtotal"
    ws["D7"] = "=SUM({{items.range.D}})"
    ws["A8"] = "Tax"
    ws["D8"] = "=SUM({{items.range.D}})*0.1"
    ws["A9"] = "Lines"
    ws["D9"] = "={{items.count}}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_per_row_formula_uses_current_row(app, db):
    po = _seed_po(db)
    rendered = render_po_xlsx(po, _build_formula_template_bytes())
    ws = load_workbook(io.BytesIO(rendered)).active
    # Open marker at row 3, single template row at 4, close at 5.
    # Two PO lines -> rows 3 and 4 hold the expanded lines.
    assert ws.cell(row=3, column=4).value == "=B3*C3"
    assert ws.cell(row=4, column=4).value == "=B4*C4"


def test_items_range_resolves_to_expanded_range(app, db):
    po = _seed_po(db)
    rendered = render_po_xlsx(po, _build_formula_template_bytes())
    ws = load_workbook(io.BytesIO(rendered)).active
    # Subtotal/Tax/Lines were at rows 7/8/9 in the template. After expansion
    # (3 marker+template rows replaced by 2 line rows, net -1), they sit at
    # rows 6/7/8. Easier to find by their label text.
    label_rows = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value in ("Subtotal", "Tax", "Lines"):
                label_rows[c.value] = c.row
    subtotal = ws.cell(row=label_rows["Subtotal"], column=4).value
    tax = ws.cell(row=label_rows["Tax"], column=4).value
    count = ws.cell(row=label_rows["Lines"], column=4).value
    # Items expanded into rows 3 and 4 -> range D3:D4.
    assert subtotal == "=SUM(D3:D4)"
    assert tax == "=SUM(D3:D4)*0.1"
    assert count == "=2"


def test_items_range_with_no_lines_collapses_to_zero(app, db):
    po = PurchaseOrder(po_number="PO-empty", vendor="None")
    db.session.add(po)
    db.session.commit()
    rendered = render_po_xlsx(po, _build_formula_template_bytes())
    ws = load_workbook(io.BytesIO(rendered)).active
    # No lines: items.range.D collapses to "0" so =SUM(0) stays valid.
    for row in ws.iter_rows():
        for c in row:
            if c.value == "Subtotal":
                assert ws.cell(row=c.row, column=4).value == "=SUM(0)"
            if c.value == "Tax":
                assert ws.cell(row=c.row, column=4).value == "=SUM(0)*0.1"
            if c.value == "Lines":
                assert ws.cell(row=c.row, column=4).value == "=0"


def test_sample_template_renders(app, db):
    """Render against the bundled sample template; must not throw."""
    po = _seed_po(db)
    sample = Path(__file__).resolve().parent.parent / "sample_template" / "po_template.xlsx"
    rendered = render_po_xlsx(po, sample.read_bytes())
    wb = load_workbook(io.BytesIO(rendered))
    ws = wb.active
    # Sanity check: PO number landed somewhere.
    found_po_number = False
    for row in ws.iter_rows():
        for c in row:
            if c.value == "PO-2026-001":
                found_po_number = True
    assert found_po_number, "Rendered template missing PO number"
