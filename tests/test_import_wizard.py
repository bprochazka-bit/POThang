"""Tests for the import wizard."""
from __future__ import annotations

import io
import json

import pytest

from purchasetracker.import_parsers import (
    parse_upload, suggest_mapping, _looks_like_header, _is_numeric,
)
from purchasetracker.models import Item


# ---------- Parser ----------

class TestParser:
    def test_csv_with_header(self):
        blob = b"name,vendor,price\nGPU,Newegg,1899\nTent,REI,200\n"
        r = parse_upload("test.csv", blob)
        assert r["had_header_row"] is True
        assert r["headers"] == ["name", "vendor", "price"]
        assert len(r["rows"]) == 2
        assert r["rows"][0] == {"name": "GPU", "vendor": "Newegg", "price": "1899"}

    def test_csv_without_header(self):
        blob = b"GPU,Newegg,1899\nTent,REI,200\n"
        r = parse_upload("test.csv", blob)
        assert r["had_header_row"] is False
        assert r["headers"] == ["Column 1", "Column 2", "Column 3"]
        assert len(r["rows"]) == 2

    def test_csv_with_currency_in_first_row_is_data(self):
        """A first row containing $1,899.00 should be detected as data, not header."""
        blob = b"GPU,Newegg,$1899.00\nTent,REI,$200\n"
        r = parse_upload("test.csv", blob)
        assert r["had_header_row"] is False

    def test_tsv(self):
        blob = b"name\tvendor\tprice\nGPU\tNewegg\t1899\n"
        r = parse_upload("test.tsv", blob)
        assert r["headers"] == ["name", "vendor", "price"]
        assert r["rows"][0]["vendor"] == "Newegg"

    def test_csv_sniffs_semicolon_delimiter(self):
        """European-style CSV with ; as delimiter."""
        blob = b"name;vendor;price\nGPU;Newegg;1899\n"
        r = parse_upload("test.csv", blob)
        assert r["headers"] == ["name", "vendor", "price"]

    def test_xlsx(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(["Product", "Manufacturer", "List Price"])
        ws.append(["GPU", "Newegg", 1899])
        ws.append(["Tent", "REI", 200])
        buf = io.BytesIO()
        wb.save(buf)

        r = parse_upload("test.xlsx", buf.getvalue())
        assert r["format"] == "xlsx"
        assert r["had_header_row"] is True
        assert r["headers"] == ["Product", "Manufacturer", "List Price"]
        assert len(r["rows"]) == 2

    def test_json_list_of_objects(self):
        blob = json.dumps([
            {"item": "GPU", "mfr": "Newegg", "price": 1899},
            {"item": "Tent", "mfr": "REI", "price": 200},
        ]).encode()
        r = parse_upload("test.json", blob)
        assert r["headers"] == ["item", "mfr", "price"]
        assert len(r["rows"]) == 2

    def test_json_pt_export_format(self):
        """Our own JSON export shape: {version: 1, items: [...]}."""
        blob = json.dumps({
            "version": 1,
            "items": [{"name": "X", "vendor": "Y", "qty": 1}]
        }).encode()
        r = parse_upload("export.json", blob)
        assert "name" in r["headers"]
        assert "vendor" in r["headers"]
        assert len(r["rows"]) == 1

    def test_json_with_list_values_joined(self):
        """List values (e.g. tags) are flattened to comma strings."""
        blob = json.dumps([{"name": "X", "tags": ["lab", "urgent"]}]).encode()
        r = parse_upload("test.json", blob)
        assert r["rows"][0]["tags"] == "lab, urgent"

    def test_format_sniffing_when_extension_missing(self):
        """No extension - sniff content to decide."""
        blob = b'[{"name":"X"}]'
        r = parse_upload("noext", blob)
        assert r["format"] == "json"

    def test_xlsx_sniff_by_magic_bytes(self):
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.append(["a", "b"]); ws.append(["1", "2"])
        buf = io.BytesIO(); wb.save(buf)
        r = parse_upload("noext", buf.getvalue())
        assert r["format"] == "xlsx"

    def test_empty_csv(self):
        r = parse_upload("empty.csv", b"")
        assert r["headers"] == []
        assert r["rows"] == []


class TestHeaderHeuristic:
    def test_known_tokens_detected(self):
        assert _looks_like_header(["name", "vendor", "price"])
        assert _looks_like_header(["Product", "Manufacturer", "Cost"])

    def test_numeric_row_rejected(self):
        assert not _looks_like_header(["1", "2", "3"])
        assert not _looks_like_header(["GPU", "Newegg", "1899"])

    def test_url_row_rejected(self):
        assert not _looks_like_header(["GPU", "https://newegg.example/x", "1899"])

    def test_currency_row_rejected(self):
        assert not _looks_like_header(["GPU", "Newegg", "$1,899.00"])

    def test_is_numeric_handles_currency(self):
        assert _is_numeric("$1,899.00")
        assert _is_numeric("1899")
        assert not _is_numeric("RTX 4090")


# ---------- Mapping suggestion ----------

class TestSuggestMapping:
    def test_exact_pt_field_names_match(self):
        m = suggest_mapping(["name", "vendor", "url", "qty", "unit_cost"])
        assert m["name"] == "name"
        assert m["vendor"] == "vendor"
        assert m["url"] == "url"
        assert m["qty"] == "qty"
        assert m["unit_cost"] == "unit_cost"

    def test_aliases(self):
        m = suggest_mapping(["Manufacturer", "Product", "Part No", "List Price"])
        assert m["vendor"] == "Manufacturer"
        assert m["name"] == "Product"
        assert m["vendor_sku"] == "Part No"
        assert m["unit_cost"] == "List Price"

    def test_no_match_for_unknown_headers(self):
        m = suggest_mapping(["foo", "bar", "baz"])
        assert m == {}

    def test_each_source_used_at_most_once(self):
        """Only one PT field can claim a given source header.

        With the alias rules, when only 'description' is present it's claimed
        by 'name' (since 'description' is a legacy name alias for backward
        compat with v3/v4 exports), and then no other PT field gets it.
        """
        m = suggest_mapping(["description"])
        # 'description' should be claimed by exactly one PT field
        users = [k for k, v in m.items() if v == "description"]
        assert len(users) == 1, f"Expected 1 user, got {users}"


# ---------- Wizard endpoints ----------

class TestWizardFlow:
    def _upload(self, client, content: bytes, filename: str):
        return client.post(
            "/io/import-wizard",
            data={"file": (io.BytesIO(content), filename)},
            content_type="multipart/form-data",
            follow_redirects=False,
        )

    def test_upload_redirects_to_map(self, client, db):
        resp = self._upload(client, b"name,vendor\nGPU,Newegg\n", "x.csv")
        assert resp.status_code == 302
        assert "/import-wizard/" in resp.headers["Location"]
        assert "/map" in resp.headers["Location"]

    def test_upload_rejects_empty_file(self, client, db):
        resp = self._upload(client, b"", "x.csv")
        assert resp.status_code == 302
        assert resp.headers["Location"].endswith("/import-wizard")

    def test_map_page_shows_preview_and_suggestions(self, client, db):
        self._upload(client, b"Product,Manufacturer,Cost\nGPU,Newegg,1899\n",
                     "catalog.csv")
        # Get the redirect target
        with client.session_transaction() as s:
            sid = s.get("import_staging_uuid")
        resp = client.get(f"/io/import-wizard/{sid}/map")
        assert resp.status_code == 200
        body = resp.data.decode()
        # Headers shown
        assert "Product" in body
        assert "Manufacturer" in body
        # Pre-selected suggestions: name -> Product, vendor -> Manufacturer
        # Just look for the right options being marked selected
        assert 'value="Product" selected' in body
        assert 'value="Manufacturer" selected' in body

    def test_full_happy_path(self, client, db):
        csv_blob = (
            b"Product,Manufacturer,Cost,URL,Quantity,Notes\n"
            b"GPU,Newegg,1899,https://newegg.example/gpu,1,Founders\n"
            b"Tent,REI,200,https://rei.example/tent,2,8-person\n"
        )
        self._upload(client, csv_blob, "catalog.csv")
        with client.session_transaction() as s:
            sid = s["import_staging_uuid"]

        # Submit mapping
        resp = client.post(f"/io/import-wizard/{sid}/map", data={
            "map_name": "Product",
            "map_vendor": "Manufacturer",
            "map_unit_cost": "Cost",
            "map_url": "URL",
            "map_qty": "Quantity",
            "map_notes": "Notes",
            "map_description": "__constant__",
            "const_description": "Imported from vendor catalog",
        }, follow_redirects=False)
        assert resp.status_code == 302
        assert "/review" in resp.headers["Location"]

        # Review page renders both rows
        resp = client.get(f"/io/import-wizard/{sid}/review")
        body = resp.data.decode()
        assert "GPU" in body
        assert "Tent" in body
        assert "Newegg" in body
        assert "1899" in body
        # Constant should be applied
        assert "Imported from vendor catalog" in body

        # Commit
        resp = client.post(f"/io/import-wizard/{sid}/review", data={
            "mode": "commit",
            "edit_0_name": "GPU",
            "edit_0_description": "Imported from vendor catalog",
            "edit_0_vendor": "Newegg",
            "edit_0_url": "https://newegg.example/gpu",
            "edit_0_qty": "1",
            "edit_0_unit_cost": "1899",
            "edit_0_notes": "Founders",
            "edit_1_name": "Tent",
            "edit_1_description": "Imported from vendor catalog",
            "edit_1_vendor": "REI",
            "edit_1_url": "https://rei.example/tent",
            "edit_1_qty": "2",
            "edit_1_unit_cost": "200",
            "edit_1_notes": "8-person",
        }, follow_redirects=False)
        assert resp.status_code == 302
        assert resp.headers["Location"].endswith("/items/")

        items = db.session.query(Item).order_by(Item.name).all()
        assert len(items) == 2
        gpu = next(i for i in items if i.name == "GPU")
        assert gpu.vendor == "Newegg"
        assert gpu.unit_cost == 1899.0
        assert gpu.url == "https://newegg.example/gpu"
        assert gpu.is_complete is True

    def test_inline_edit_overrides_mapped_value(self, client, db):
        """Editing a cell on the review page should be reflected in commit."""
        self._upload(client, b"name,vendor\nGPU,Newegg\n", "x.csv")
        with client.session_transaction() as s:
            sid = s["import_staging_uuid"]
        client.post(f"/io/import-wizard/{sid}/map", data={
            "map_name": "name", "map_vendor": "vendor",
        })
        # Commit with an edit that overrides vendor
        client.post(f"/io/import-wizard/{sid}/review", data={
            "mode": "commit",
            "edit_0_name": "GPU",
            "edit_0_vendor": "MicroCenter",  # changed!
            "edit_0_url": "https://x",
            "edit_0_qty": "1",
            "edit_0_unit_cost": "100",
            "edit_0_description": "card",
        })
        item = db.session.query(Item).one()
        assert item.vendor == "MicroCenter"

    def test_constant_value_applied_to_all_rows(self, client, db):
        csv_blob = b"name,price\nGPU,1899\nTent,200\n"
        self._upload(client, csv_blob, "x.csv")
        with client.session_transaction() as s:
            sid = s["import_staging_uuid"]
        client.post(f"/io/import-wizard/{sid}/map", data={
            "map_name": "name",
            "map_unit_cost": "price",
            "map_vendor": "__constant__",
            "const_vendor": "Acme",
        })
        client.post(f"/io/import-wizard/{sid}/review", data={
            "mode": "commit",
            "edit_0_name": "GPU", "edit_0_vendor": "Acme",
            "edit_0_unit_cost": "1899", "edit_0_url": "u",
            "edit_0_qty": "1", "edit_0_description": "d",
            "edit_1_name": "Tent", "edit_1_vendor": "Acme",
            "edit_1_unit_cost": "200", "edit_1_url": "u",
            "edit_1_qty": "1", "edit_1_description": "d",
        })
        items = db.session.query(Item).all()
        assert all(i.vendor == "Acme" for i in items)

    def test_map_requires_name(self, client, db):
        """Cannot proceed without a name mapping (or constant)."""
        self._upload(client, b"vendor\nNewegg\n", "x.csv")
        with client.session_transaction() as s:
            sid = s["import_staging_uuid"]
        resp = client.post(f"/io/import-wizard/{sid}/map", data={
            "map_vendor": "vendor",
            # no map_name
        }, follow_redirects=False)
        assert resp.status_code == 302
        assert "/map" in resp.headers["Location"]  # bounced back to mapper

    def test_save_edits_does_not_commit(self, client, db):
        self._upload(client, b"name\nGPU\n", "x.csv")
        with client.session_transaction() as s:
            sid = s["import_staging_uuid"]
        client.post(f"/io/import-wizard/{sid}/map", data={"map_name": "name"})
        resp = client.post(f"/io/import-wizard/{sid}/review", data={
            "mode": "save_edits",
            "edit_0_name": "GPU edited",
        }, follow_redirects=False)
        # No items should have been created
        assert db.session.query(Item).count() == 0
        # Should still be on the review page
        assert "/review" in resp.headers["Location"]

    def test_cancel_clears_staging(self, client, db):
        self._upload(client, b"name\nGPU\n", "x.csv")
        with client.session_transaction() as s:
            sid = s["import_staging_uuid"]
        resp = client.post(f"/io/import-wizard/{sid}/cancel",
                           follow_redirects=False)
        assert resp.status_code == 302
        # Subsequent map access should redirect away
        resp = client.get(f"/io/import-wizard/{sid}/map", follow_redirects=False)
        assert resp.status_code == 302  # bounced back to upload

    def test_xlsx_upload_through_wizard(self, client, db):
        from openpyxl import Workbook
        wb = Workbook(); ws = wb.active
        ws.append(["Item Name", "Brand", "Each", "Quantity", "Description", "Web Address"])
        ws.append(["Widget", "WidgetCo", 9.99, 5, "Stainless steel", "https://w.example"])
        buf = io.BytesIO(); wb.save(buf)
        self._upload(client, buf.getvalue(), "catalog.xlsx")
        with client.session_transaction() as s:
            sid = s["import_staging_uuid"]
        # Check map page shows xlsx headers
        resp = client.get(f"/io/import-wizard/{sid}/map")
        body = resp.data.decode()
        assert "Item Name" in body
        assert "Brand" in body
        # Auto-suggestion: Item Name -> name, Brand -> vendor, Each -> unit_cost
        assert 'value="Item Name" selected' in body
        assert 'value="Brand" selected' in body

    def test_re_import_updates_existing_item(self, client, db):
        """Match by (name, vendor, model) - second import updates instead of duplicating."""
        # First import
        self._upload(client, b"name,vendor,price\nGPU,Newegg,1899\n", "x.csv")
        with client.session_transaction() as s:
            sid = s["import_staging_uuid"]
        client.post(f"/io/import-wizard/{sid}/map", data={
            "map_name": "name", "map_vendor": "vendor", "map_unit_cost": "price",
        })
        client.post(f"/io/import-wizard/{sid}/review", data={
            "mode": "commit",
            "edit_0_name": "GPU", "edit_0_vendor": "Newegg",
            "edit_0_unit_cost": "1899", "edit_0_url": "https://x",
            "edit_0_qty": "1", "edit_0_description": "d",
        })
        assert db.session.query(Item).count() == 1

        # Second import with updated price
        self._upload(client, b"name,vendor,price\nGPU,Newegg,1799\n", "x2.csv")
        with client.session_transaction() as s:
            sid2 = s["import_staging_uuid"]
        client.post(f"/io/import-wizard/{sid2}/map", data={
            "map_name": "name", "map_vendor": "vendor", "map_unit_cost": "price",
        })
        client.post(f"/io/import-wizard/{sid2}/review", data={
            "mode": "commit",
            "edit_0_name": "GPU", "edit_0_vendor": "Newegg",
            "edit_0_unit_cost": "1799", "edit_0_url": "https://x",
            "edit_0_qty": "1", "edit_0_description": "d",
        })
        # Still one item, with updated price
        items = db.session.query(Item).all()
        assert len(items) == 1
        assert items[0].unit_cost == 1799.0
