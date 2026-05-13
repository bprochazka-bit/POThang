"""Tests for the server-side PO document archive (revision tracking)."""
from __future__ import annotations

import io
from pathlib import Path

from openpyxl import Workbook

from purchasetracker.models import Item, PODocument, POLine, PurchaseOrder
from purchasetracker.services import (
    next_po_revision, po_document_path, store_po_document,
)


def _build_template_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "PO: {{po_number}}"
    ws["A2"] = "Revision: {{revision}}"
    ws["A4"] = "{{#items}}"
    ws["A5"] = "{{item.index}}"
    ws["B5"] = "{{item.name}}"
    ws["A6"] = "{{/items}}"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _seed_po_with_lines(db) -> PurchaseOrder:
    item = Item(name="Widget", description="W", qty=2, unit_cost=5.0,
                vendor="V", url="http://v")
    db.session.add(item)
    db.session.flush()
    po = PurchaseOrder(po_number="PO-DOC-1", vendor="V")
    db.session.add(po)
    db.session.flush()
    db.session.add(POLine(po_id=po.id, item_id=item.id, qty=2, unit_cost=5.0))
    db.session.commit()
    return po


def test_next_po_revision_starts_at_one(app, db):
    po = _seed_po_with_lines(db)
    assert next_po_revision(po.id) == 1


def test_store_po_document_assigns_increasing_revisions(app, db):
    po = _seed_po_with_lines(db)
    doc1 = store_po_document(po, b"first", template_name="t.xlsx")
    db.session.commit()
    doc2 = store_po_document(po, b"second", template_name="t.xlsx")
    db.session.commit()
    assert doc1.revision == 1
    assert doc2.revision == 2


def test_store_po_document_writes_blob_to_disk(app, db):
    po = _seed_po_with_lines(db)
    doc = store_po_document(po, b"hello world")
    db.session.commit()
    path = po_document_path(doc)
    assert path.exists()
    assert path.read_bytes() == b"hello world"


def test_store_po_document_filename_includes_revision(app, db):
    po = _seed_po_with_lines(db)
    doc = store_po_document(po, b"x")
    db.session.commit()
    assert doc.original_filename == "PO-DOC-1-rev1.xlsx"


def test_identical_content_dedupes_on_disk(app, db):
    po = _seed_po_with_lines(db)
    doc1 = store_po_document(po, b"same bytes")
    doc2 = store_po_document(po, b"same bytes")
    db.session.commit()
    # Different revisions but same hash → same on-disk path, single blob.
    assert doc1.sha256 == doc2.sha256
    assert doc1.revision != doc2.revision
    assert po_document_path(doc1) == po_document_path(doc2)


def test_render_xlsx_route_archives_revision(client, db):
    """Hitting /<po_id>/render saves a PODocument and returns the file."""
    po = _seed_po_with_lines(db)

    tpl = _build_template_bytes()
    resp = client.post(
        f"/pos/{po.id}/render",
        data={
            "template_name": "__upload__",
            "template": (io.BytesIO(tpl), "tmpl.xlsx"),
        },
        content_type="multipart/form-data",
    )
    assert resp.status_code == 200
    assert resp.headers["Content-Disposition"].startswith("attachment")
    # Filename should encode revision 1.
    assert "rev1" in resp.headers["Content-Disposition"]

    docs = db.session.query(PODocument).filter_by(po_id=po.id).all()
    assert len(docs) == 1
    assert docs[0].revision == 1
    assert docs[0].template_name and docs[0].template_name.startswith("upload:")


def test_repeat_render_creates_new_revision(client, db):
    po = _seed_po_with_lines(db)
    tpl = _build_template_bytes()
    for _ in range(3):
        client.post(
            f"/pos/{po.id}/render",
            data={
                "template_name": "__upload__",
                "template": (io.BytesIO(tpl), "tmpl.xlsx"),
            },
            content_type="multipart/form-data",
        )
    docs = (db.session.query(PODocument)
            .filter_by(po_id=po.id)
            .order_by(PODocument.revision).all())
    assert [d.revision for d in docs] == [1, 2, 3]


def test_archived_doc_is_downloadable(client, db):
    po = _seed_po_with_lines(db)
    tpl = _build_template_bytes()
    client.post(
        f"/pos/{po.id}/render",
        data={
            "template_name": "__upload__",
            "template": (io.BytesIO(tpl), "tmpl.xlsx"),
        },
        content_type="multipart/form-data",
    )
    doc = db.session.query(PODocument).filter_by(po_id=po.id).one()
    resp = client.get(f"/pos/{po.id}/documents/{doc.id}/download")
    assert resp.status_code == 200
    # Should be the same bytes as on disk.
    assert resp.data == po_document_path(doc).read_bytes()


def test_render_substitutes_revision_placeholder(client, db):
    """Templates containing {{revision}} get the assigned revision number."""
    from openpyxl import load_workbook
    po = _seed_po_with_lines(db)
    tpl = _build_template_bytes()
    # First render → revision 1
    resp = client.post(
        f"/pos/{po.id}/render",
        data={
            "template_name": "__upload__",
            "template": (io.BytesIO(tpl), "tmpl.xlsx"),
        },
        content_type="multipart/form-data",
    )
    wb = load_workbook(io.BytesIO(resp.data))
    ws = wb.active
    assert ws["A2"].value == "Revision: 1"


def test_delete_document_removes_row(client, db):
    po = _seed_po_with_lines(db)
    tpl = _build_template_bytes()
    client.post(
        f"/pos/{po.id}/render",
        data={
            "template_name": "__upload__",
            "template": (io.BytesIO(tpl), "tmpl.xlsx"),
        },
        content_type="multipart/form-data",
    )
    doc = db.session.query(PODocument).filter_by(po_id=po.id).one()
    blob = po_document_path(doc)
    assert blob.exists()
    resp = client.post(f"/pos/{po.id}/documents/{doc.id}/delete",
                       follow_redirects=True)
    assert resp.status_code == 200
    assert db.session.query(PODocument).filter_by(po_id=po.id).count() == 0
    # Blob is gone too (nothing else referenced this hash).
    assert not blob.exists()


def test_delete_document_does_not_renumber(client, db):
    """Deleting rev 1 leaves later revisions with their original numbers."""
    po = _seed_po_with_lines(db)
    tpl = _build_template_bytes()
    for _ in range(3):
        client.post(
            f"/pos/{po.id}/render",
            data={
                "template_name": "__upload__",
                "template": (io.BytesIO(tpl), "tmpl.xlsx"),
            },
            content_type="multipart/form-data",
        )
    rev1 = (db.session.query(PODocument)
            .filter_by(po_id=po.id, revision=1).one())
    client.post(f"/pos/{po.id}/documents/{rev1.id}/delete",
                follow_redirects=True)
    remaining = (db.session.query(PODocument)
                 .filter_by(po_id=po.id)
                 .order_by(PODocument.revision).all())
    assert [d.revision for d in remaining] == [2, 3]
    # Next render bumps past the deleted rev1.
    client.post(
        f"/pos/{po.id}/render",
        data={
            "template_name": "__upload__",
            "template": (io.BytesIO(tpl), "tmpl.xlsx"),
        },
        content_type="multipart/form-data",
    )
    revs = sorted(d.revision for d in db.session.query(PODocument)
                  .filter_by(po_id=po.id).all())
    assert revs == [2, 3, 4]


def test_documents_listed_on_detail_page(client, db):
    po = _seed_po_with_lines(db)
    tpl = _build_template_bytes()
    client.post(
        f"/pos/{po.id}/render",
        data={
            "template_name": "__upload__",
            "template": (io.BytesIO(tpl), "tmpl.xlsx"),
        },
        content_type="multipart/form-data",
    )
    resp = client.get(f"/pos/{po.id}")
    body = resp.data.decode()
    assert "Saved revisions" in body
    assert "rev 1" in body


def test_download_rejects_doc_from_wrong_po(client, db):
    po1 = _seed_po_with_lines(db)
    # Build a second PO
    po2 = PurchaseOrder(po_number="PO-OTHER")
    db.session.add(po2)
    db.session.commit()
    tpl = _build_template_bytes()
    client.post(
        f"/pos/{po1.id}/render",
        data={
            "template_name": "__upload__",
            "template": (io.BytesIO(tpl), "tmpl.xlsx"),
        },
        content_type="multipart/form-data",
    )
    doc = db.session.query(PODocument).filter_by(po_id=po1.id).one()
    # Asking via the wrong PO id should 404.
    resp = client.get(f"/pos/{po2.id}/documents/{doc.id}/download")
    assert resp.status_code == 404
