"""Tests for the stable per-PO line_no introduced for consistent numbering."""
from __future__ import annotations

import io

from openpyxl import Workbook, load_workbook

from purchasetracker.models import Item, POLine, PurchaseOrder
from purchasetracker.services import move_po_line, render_po_xlsx


def _seed_complete_item(db, name: str, vendor: str = "V", qty: int = 10):
    item = Item(
        name=name, description=f"{name} desc", qty=qty, unit_cost=1.0,
        vendor=vendor, url=f"http://example/{name.lower()}",
    )
    db.session.add(item)
    db.session.flush()
    return item


def test_line_no_auto_assigned_on_direct_insert(app, db):
    po = PurchaseOrder(po_number="PO-LN-1")
    db.session.add(po)
    db.session.flush()
    a = _seed_complete_item(db, "Alpha")
    b = _seed_complete_item(db, "Bravo")
    c = _seed_complete_item(db, "Charlie")
    db.session.add_all([
        POLine(po_id=po.id, item_id=a.id, qty=1, unit_cost=1.0),
        POLine(po_id=po.id, item_id=b.id, qty=1, unit_cost=1.0),
        POLine(po_id=po.id, item_id=c.id, qty=1, unit_cost=1.0),
    ])
    db.session.commit()
    nos = sorted(l.line_no for l in po.lines)
    assert nos == [1, 2, 3]


def test_line_no_continues_after_existing_rows(app, db):
    """Adding a new line after some exist picks max(line_no)+1, not 1."""
    po = PurchaseOrder(po_number="PO-LN-2")
    db.session.add(po)
    db.session.flush()
    a = _seed_complete_item(db, "Alpha")
    b = _seed_complete_item(db, "Bravo")
    db.session.add(POLine(po_id=po.id, item_id=a.id, qty=1, unit_cost=1.0))
    db.session.commit()
    # Second flush, separate transaction
    db.session.add(POLine(po_id=po.id, item_id=b.id, qty=1, unit_cost=1.0))
    db.session.commit()
    nos = sorted(l.line_no for l in po.lines)
    assert nos == [1, 2]


def test_line_no_independent_per_po(app, db):
    po1 = PurchaseOrder(po_number="PO-LN-A")
    po2 = PurchaseOrder(po_number="PO-LN-B")
    db.session.add_all([po1, po2])
    db.session.flush()
    a = _seed_complete_item(db, "Alpha")
    db.session.add_all([
        POLine(po_id=po1.id, item_id=a.id, qty=1, unit_cost=1.0),
        POLine(po_id=po2.id, item_id=a.id, qty=1, unit_cost=1.0),
    ])
    db.session.commit()
    # Each PO starts its own line numbering at 1.
    assert po1.lines[0].line_no == 1
    assert po2.lines[0].line_no == 1


def test_line_no_survives_blueprint_add(client, db):
    """Adding via the pos blueprint also assigns line_no."""
    a = _seed_complete_item(db, "Alpha", vendor="X")
    b = _seed_complete_item(db, "Bravo", vendor="Y")
    db.session.commit()
    client.post("/pos/new", data={"po_number": "PO-BP"}, follow_redirects=True)
    po = db.session.query(PurchaseOrder).filter_by(po_number="PO-BP").one()
    client.post(f"/pos/{po.id}/lines/add",
                data={"item_id": a.id, "qty": "1"}, follow_redirects=True)
    client.post(f"/pos/{po.id}/lines/add",
                data={"item_id": b.id, "qty": "1"}, follow_redirects=True)
    lines = db.session.query(POLine).filter_by(po_id=po.id).order_by(POLine.id).all()
    assert [l.line_no for l in lines] == [1, 2]


def test_render_uses_line_no_not_alphabetical(app, db):
    """Rendered xlsx prints lines in line_no order, not vendor/name order."""
    po = PurchaseOrder(po_number="PO-RENDER-ORDER", vendor="Acme")
    db.session.add(po)
    db.session.flush()
    # Insert Z-named first, then A-named. The old code would have sorted by
    # name and put Alpha first; the new code preserves insertion order via
    # line_no.
    z = _seed_complete_item(db, "Zeta")
    a = _seed_complete_item(db, "Alpha")
    db.session.add_all([
        POLine(po_id=po.id, item_id=z.id, qty=1, unit_cost=1.0),
        POLine(po_id=po.id, item_id=a.id, qty=1, unit_cost=1.0),
    ])
    db.session.commit()

    wb = Workbook()
    ws = wb.active
    ws["A1"] = "{{#items}}"
    ws["A2"] = "{{item.index}}"
    ws["B2"] = "{{item.name}}"
    ws["A3"] = "{{/items}}"
    buf = io.BytesIO()
    wb.save(buf)

    rendered = render_po_xlsx(po, buf.getvalue())
    out = load_workbook(io.BytesIO(rendered)).active
    # First rendered row is the one inserted first (Zeta), with index 1.
    assert out.cell(row=1, column=1).value == 1
    assert out.cell(row=1, column=2).value == "Zeta"
    assert out.cell(row=2, column=1).value == 2
    assert out.cell(row=2, column=2).value == "Alpha"


def test_move_line_swaps_neighbouring_line_nos(app, db):
    po = PurchaseOrder(po_number="PO-MOVE")
    db.session.add(po)
    db.session.flush()
    a = _seed_complete_item(db, "Alpha")
    b = _seed_complete_item(db, "Bravo")
    c = _seed_complete_item(db, "Charlie")
    db.session.add_all([
        POLine(po_id=po.id, item_id=a.id, qty=1, unit_cost=1.0),
        POLine(po_id=po.id, item_id=b.id, qty=1, unit_cost=1.0),
        POLine(po_id=po.id, item_id=c.id, qty=1, unit_cost=1.0),
    ])
    db.session.commit()

    line_b = next(l for l in po.lines if l.line_no == 2)
    assert move_po_line(line_b, -1) is True
    db.session.commit()
    # Bravo now line_no 1, Alpha line_no 2, Charlie stays at 3.
    nos = {l.item.name: l.line_no for l in po.lines}
    assert nos == {"Alpha": 2, "Bravo": 1, "Charlie": 3}


def test_move_line_at_edge_is_noop(app, db):
    po = PurchaseOrder(po_number="PO-MOVE-EDGE")
    db.session.add(po)
    db.session.flush()
    a = _seed_complete_item(db, "Alpha")
    b = _seed_complete_item(db, "Bravo")
    db.session.add_all([
        POLine(po_id=po.id, item_id=a.id, qty=1, unit_cost=1.0),
        POLine(po_id=po.id, item_id=b.id, qty=1, unit_cost=1.0),
    ])
    db.session.commit()
    line_a = next(l for l in po.lines if l.line_no == 1)
    assert move_po_line(line_a, -1) is False
    line_b = next(l for l in po.lines if l.line_no == 2)
    assert move_po_line(line_b, 1) is False


def test_move_line_via_blueprint(client, db):
    a = _seed_complete_item(db, "Alpha")
    b = _seed_complete_item(db, "Bravo")
    db.session.commit()
    client.post("/pos/new", data={"po_number": "PO-MOVE-BP"}, follow_redirects=True)
    po = db.session.query(PurchaseOrder).filter_by(po_number="PO-MOVE-BP").one()
    client.post(f"/pos/{po.id}/lines/add",
                data={"item_id": a.id, "qty": "1"}, follow_redirects=True)
    client.post(f"/pos/{po.id}/lines/add",
                data={"item_id": b.id, "qty": "1"}, follow_redirects=True)
    line_b = db.session.query(POLine).filter_by(po_id=po.id, item_id=b.id).one()
    resp = client.post(f"/pos/lines/{line_b.id}/move",
                       data={"dir": "up"}, follow_redirects=True)
    assert resp.status_code == 200
    db.session.refresh(line_b)
    assert line_b.line_no == 1


def test_line_no_unchanged_when_qty_updated(client, db):
    """Editing qty must not perturb line numbering."""
    a = _seed_complete_item(db, "Alpha", qty=10)
    db.session.commit()
    client.post("/pos/new", data={"po_number": "PO-STBL"}, follow_redirects=True)
    po = db.session.query(PurchaseOrder).filter_by(po_number="PO-STBL").one()
    client.post(f"/pos/{po.id}/lines/add",
                data={"item_id": a.id, "qty": "2"}, follow_redirects=True)
    line = db.session.query(POLine).filter_by(po_id=po.id).one()
    original = line.line_no
    client.post(f"/pos/lines/{line.id}/qty",
                data={"qty": "5"}, follow_redirects=True)
    db.session.refresh(line)
    assert line.line_no == original
