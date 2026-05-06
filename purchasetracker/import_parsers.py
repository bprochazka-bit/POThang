"""
Source-document parsers for the import wizard.

Each parser returns a uniform structure:

    {
        "headers": ["Name", "Vendor", ...],   # list of column labels
        "rows":    [{"Name": "...", "Vendor": "..."}, ...],
        "format":  "csv" | "tsv" | "xlsx" | "json",
        "had_header_row": True | False,       # we tried to detect it
    }

Headers are kept as the source supplied them (case preserved) so the user
sees their own column names in the mapper UI.

For headerless input we synthesize column labels "Column 1", "Column 2", ...
"""
from __future__ import annotations

import csv
import io
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


# Headers we recognise as candidates for the user's data. Used by the
# heuristic that decides whether row 1 is a header row, and by the auto-
# mapping suggestions in the wizard.
_KNOWN_HEADER_TOKENS = {
    "name", "title", "item", "description", "desc", "details",
    "vendor", "supplier", "manufacturer", "mfr", "mfg", "brand",
    "model", "model number", "part", "part number", "p/n", "pn",
    "sku", "vendor sku", "stock number", "stock no",
    "url", "link", "website", "product url", "page",
    "qty", "quantity", "count", "units",
    "cost", "price", "unit cost", "unit price", "list price", "msrp",
    "tags", "category", "categories",
    "notes", "comment", "comments", "remark", "remarks",
}


# ---------- public API ----------

def parse_upload(filename: str, blob: bytes) -> dict:
    """Auto-detect format from filename and parse. Returns dict shape above."""
    ext = (Path(filename).suffix or "").lower().lstrip(".")
    if ext in ("xlsx", "xlsm"):
        return _parse_xlsx(blob)
    if ext == "json":
        return _parse_json(blob)
    if ext == "tsv":
        return _parse_csv_like(blob, force_delim="\t", fmt="tsv")
    if ext == "csv":
        return _parse_csv_like(blob, fmt="csv")
    # Fallback: try to sniff from contents.
    sniffed = _sniff_format(blob)
    if sniffed == "json":
        return _parse_json(blob)
    if sniffed == "xlsx":
        return _parse_xlsx(blob)
    return _parse_csv_like(blob, fmt=sniffed or "csv")


def _sniff_format(blob: bytes) -> str:
    head = blob[:4]
    if head[:2] == b"PK":
        return "xlsx"
    # Quick text peek
    try:
        text = blob[:4096].decode("utf-8-sig")
    except UnicodeDecodeError:
        return "csv"
    stripped = text.lstrip()
    if stripped.startswith("{") or stripped.startswith("["):
        return "json"
    if "\t" in stripped.split("\n", 1)[0]:
        return "tsv"
    return "csv"


# ---------- CSV / TSV ----------

def _parse_csv_like(blob: bytes, force_delim: str | None = None,
                    fmt: str = "csv") -> dict:
    text = blob.decode("utf-8-sig", errors="replace")
    delim = force_delim
    if delim is None:
        try:
            dialect = csv.Sniffer().sniff(text[:4096], delimiters=",\t;|")
            delim = dialect.delimiter
        except csv.Error:
            delim = ","

    reader = csv.reader(io.StringIO(text), delimiter=delim)
    raw_rows = [r for r in reader if any(cell.strip() for cell in r)]
    if not raw_rows:
        return {"headers": [], "rows": [], "format": fmt, "had_header_row": False}

    first = raw_rows[0]
    had_header = _looks_like_header(first)
    if had_header:
        headers = [_clean_header(c) for c in first]
        body = raw_rows[1:]
    else:
        headers = [f"Column {i+1}" for i in range(len(first))]
        body = raw_rows

    rows = []
    for r in body:
        # Pad short rows; clip long ones.
        cells = list(r) + [""] * max(0, len(headers) - len(r))
        cells = cells[:len(headers)]
        rows.append({h: (c if c is not None else "") for h, c in zip(headers, cells)})

    # Deduplicate identical headers (common in xlsx exports)
    headers = _dedup(headers)
    return {"headers": headers, "rows": rows, "format": fmt,
            "had_header_row": had_header}


# ---------- xlsx ----------

def _parse_xlsx(blob: bytes) -> dict:
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    ws = wb.active
    raw_rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if v is None else str(v) for v in row]
        if any(c.strip() for c in cells):
            raw_rows.append(cells)
        if len(raw_rows) > 10000:
            break  # safety cap

    if not raw_rows:
        return {"headers": [], "rows": [], "format": "xlsx", "had_header_row": False}

    first = raw_rows[0]
    had_header = _looks_like_header(first)
    if had_header:
        headers = [_clean_header(c) for c in first]
        body = raw_rows[1:]
    else:
        headers = [f"Column {i+1}" for i in range(len(first))]
        body = raw_rows

    headers = _dedup(headers)
    rows = []
    for r in body:
        cells = list(r) + [""] * max(0, len(headers) - len(r))
        cells = cells[:len(headers)]
        rows.append({h: c for h, c in zip(headers, cells)})
    return {"headers": headers, "rows": rows, "format": "xlsx",
            "had_header_row": had_header}


# ---------- JSON ----------

def _parse_json(blob: bytes) -> dict:
    text = blob.decode("utf-8-sig", errors="replace")
    data = json.loads(text)

    # Accept either a list of objects, or a PT-style export
    # ({"items": [...]}) so the wizard works with our own JSON output too.
    if isinstance(data, dict) and isinstance(data.get("items"), list):
        records = data["items"]
    elif isinstance(data, list):
        records = data
    else:
        raise ValueError(
            "JSON must be either a list of objects or "
            "{\"items\": [...]} as produced by the export feature."
        )

    if not records:
        return {"headers": [], "rows": [], "format": "json", "had_header_row": True}

    # Headers = union of keys, in first-seen order.
    seen = []
    for rec in records:
        if not isinstance(rec, dict):
            continue
        for k in rec.keys():
            if k not in seen:
                seen.append(k)

    rows = []
    for rec in records:
        if not isinstance(rec, dict):
            continue
        out = {}
        for h in seen:
            v = rec.get(h)
            if isinstance(v, list):
                out[h] = ", ".join(str(x) for x in v)
            elif v is None:
                out[h] = ""
            else:
                out[h] = str(v)
        rows.append(out)

    return {"headers": seen, "rows": rows, "format": "json",
            "had_header_row": True}


# ---------- helpers ----------

def _clean_header(s: str) -> str:
    s = (s or "").strip()
    return s if s else "Column"


def _dedup(headers: list[str]) -> list[str]:
    """Append a counter to repeat header names so they're unique."""
    seen: dict[str, int] = {}
    out = []
    for h in headers:
        if h not in seen:
            seen[h] = 1
            out.append(h)
        else:
            seen[h] += 1
            out.append(f"{h} ({seen[h]})")
    return out


def _looks_like_header(row: list[str]) -> bool:
    """Heuristic: does this row look like column headers rather than data?

    Signals (any one of these is enough):
      - Most cells contain at least one alpha character AND no purely numeric
        values (numeric values strongly suggest a data row).
      - At least one cell matches a known header token.
      - Cells are short (< 40 chars on average) and don't contain http://...
    """
    cells = [c.strip() for c in row if c is not None]
    if not cells:
        return False

    # Strong negative: any cell looks like a number/currency/url/email
    for c in cells:
        if _is_numeric(c) or c.lower().startswith(("http://", "https://")) \
           or "@" in c:
            return False

    # Strong positive: any cell matches a known header token
    norm = [_normalize_header(c) for c in cells]
    if any(n in _KNOWN_HEADER_TOKENS for n in norm):
        return True

    # Soft positive: short, alphabetic-looking labels
    avg_len = sum(len(c) for c in cells) / len(cells)
    alpha_frac = sum(1 for c in cells if any(ch.isalpha() for ch in c)) / len(cells)
    return avg_len < 40 and alpha_frac > 0.7


def _is_numeric(s: str) -> bool:
    if not s:
        return False
    s = s.strip().lstrip("$").replace(",", "")
    try:
        float(s)
        return True
    except ValueError:
        return False


def _normalize_header(s: str) -> str:
    return re.sub(r"[^a-z0-9 ]", "", (s or "").lower()).strip()


# ---------- mapping suggestion ----------

# Each PT field maps to a list of header tokens we'll auto-match.
PT_FIELD_ALIASES: dict[str, list[str]] = {
    "name":        ["name", "title", "item", "item name", "product",
                    "product name", "description", "desc"],
    "description": ["description", "desc", "details", "long description",
                    "summary", "specifications", "specs"],
    "vendor":      ["vendor", "supplier", "manufacturer", "mfr", "mfg",
                    "brand", "seller", "source"],
    "model":       ["model", "model number", "model no", "model #"],
    "vendor_sku":  ["sku", "vendor sku", "part", "part number", "p/n", "pn",
                    "stock number", "stock no", "item number", "id"],
    "url":         ["url", "link", "website", "product url", "product link",
                    "page", "web", "web address"],
    "qty":         ["qty", "quantity", "count", "units", "amount"],
    "unit_cost":   ["unit cost", "unit price", "cost", "price", "list price",
                    "msrp", "each", "ea", "price each"],
    "tags":        ["tags", "tag", "category", "categories", "type", "group"],
    "notes":       ["notes", "note", "comment", "comments", "remark",
                    "remarks", "memo"],
}


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    """Best-guess mapping {pt_field: source_header}.

    For each PT field, find the header whose normalized text matches one of
    its aliases. Prefer exact matches over substring matches. If two PT
    fields would claim the same source header (e.g. only one column called
    "description" exists, both `name` and `description` could grab it), the
    earlier-listed PT field wins and the other gets no auto-mapping.
    """
    norm_headers = {h: _normalize_header(h) for h in headers}
    used: set[str] = set()
    out: dict[str, str] = {}

    # Two passes: exact match wins over substring match.
    for pt_field, aliases in PT_FIELD_ALIASES.items():
        match = None
        # Pass 1: exact normalized match
        for h, nh in norm_headers.items():
            if h in used:
                continue
            if nh in aliases:
                match = h
                break
        if match:
            out[pt_field] = match
            used.add(match)

    for pt_field, aliases in PT_FIELD_ALIASES.items():
        if pt_field in out:
            continue
        match = None
        # Pass 2: substring match (alias appears inside header)
        for h, nh in norm_headers.items():
            if h in used:
                continue
            for alias in aliases:
                if alias in nh or nh in alias:
                    match = h
                    break
            if match:
                break
        if match:
            out[pt_field] = match
            used.add(match)

    return out
