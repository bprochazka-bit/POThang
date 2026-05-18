"""
Service-layer helpers. Anything that's more than a CRUD operation lives here.
"""
from __future__ import annotations

import colorsys
import datetime as dt
import hashlib
import io
import os
import re
import shutil
import xml.etree.ElementTree as ET
from copy import copy
from pathlib import Path
from types import SimpleNamespace
from typing import Any, BinaryIO, Iterable, Optional, Tuple

from flask import current_app
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .extensions import db
from .models import (
    Attachment, Item, PODocument, POLine, PurchaseOrder, Receipt, Tag,
)


# ---------- State recomputation ----------

def recompute_item_state(item: Item) -> None:
    """Recompute and assign item.state from current allocations and receipts.

    The item's state mirrors the most-progressed PO it sits on:

      - received  : sum(receipts) >= item.qty
      - partial   : sum(receipts) > 0
      - ordered   : sum of qty on POs in {ordered, partial, received} >= item.qty
      - approved  : item is on at least one approved-but-not-yet-placed PO,
                    OR the user manually flagged it approved.
      - requested : everything else.

    Cancelled is terminal (user-set, never overwritten).

    Draft and submitted POs intentionally do NOT advance item state - draft
    is a working state for the buyer and submitted means the PO is awaiting
    funding approval, neither of which is a real commitment - though they
    still count toward qty_on_active_pos / qty_unallocated to prevent
    double-allocating an item to two POs at once.
    """
    if item.state == "cancelled":
        return  # terminal, user-set

    qty_total = item.qty or 0
    qty_received = item.qty_received

    # Lines on POs that have actually been placed with the vendor.
    qty_placed = sum(
        line.qty for line in item.lines
        if line.po and line.po.status in ("ordered", "partial", "received")
    )
    # Any line on an approved-but-not-yet-placed PO?
    has_approved_line = any(
        line.po is not None and line.po.status == "approved"
        for line in item.lines
    )

    if qty_received >= qty_total and qty_total > 0:
        item.state = "received"
    elif qty_received > 0:
        item.state = "partial"
    elif qty_placed >= qty_total and qty_total > 0:
        item.state = "ordered"
    elif has_approved_line or item.state == "approved":
        # On an approved PO, OR user manually flagged it approved before any
        # PO existed and we haven't progressed past that yet.
        item.state = "approved"
    else:
        item.state = "requested"


# ---------- PO line numbering ----------

def next_po_line_no(po_id: int) -> int:
    """Return the next stable line_no to assign on this PO (max + 1, or 1)."""
    current = (
        db.session.query(POLine.line_no)
        .filter(POLine.po_id == po_id)
        .order_by(POLine.line_no.desc())
        .first()
    )
    if current is None or current[0] is None:
        return 1
    return int(current[0]) + 1


def _render_sort_key(line: POLine):
    """Sort key for rendering PO lines: by stable line_no, then by id.

    line_no is 1-based for new lines; legacy rows get backfilled by migration.
    The id tiebreaker keeps things deterministic if two rows somehow share a
    line_no (shouldn't happen, but cheap insurance).
    """
    return (line.line_no or 0, line.id or 0)


def renumber_po_lines(po: PurchaseOrder) -> None:
    """Compact line_no values to 1..N preserving current line_no order."""
    ordered = sorted(po.lines, key=_render_sort_key)
    for idx, line in enumerate(ordered, start=1):
        if line.line_no != idx:
            line.line_no = idx


def move_po_line(line: POLine, direction: int) -> bool:
    """Swap a line's line_no with its neighbour above (-1) or below (+1).

    Returns True if a swap happened, False if the line was already at the
    edge. Caller is responsible for committing.
    """
    if direction not in (-1, 1):
        raise ValueError("direction must be -1 or +1")
    if line.po is None:
        return False

    siblings = sorted(line.po.lines, key=_render_sort_key)
    try:
        idx = siblings.index(line)
    except ValueError:
        return False
    neighbour_idx = idx + direction
    if neighbour_idx < 0 or neighbour_idx >= len(siblings):
        return False

    neighbour = siblings[neighbour_idx]
    line.line_no, neighbour.line_no = neighbour.line_no, line.line_no
    return True


# ---------- Saved PO documents ----------

def next_po_revision(po_id: int) -> int:
    """Return the next revision number to assign for documents on this PO."""
    current = (
        db.session.query(PODocument.revision)
        .filter(PODocument.po_id == po_id)
        .order_by(PODocument.revision.desc())
        .first()
    )
    if current is None or current[0] is None:
        return 1
    return int(current[0]) + 1


def store_po_document(po: PurchaseOrder, content: bytes,
                      template_name: Optional[str] = None,
                      generated_by: Optional[str] = None,
                      revision: Optional[int] = None,
                      mime_type: str = (
                          "application/vnd.openxmlformats-"
                          "officedocument.spreadsheetml.sheet"
                      )) -> PODocument:
    """Archive a freshly-rendered PO xlsx and return the PODocument row.

    The blob is written under uploads/<aa>/<bb>/<sha256> (same hash-tree as
    Attachment) so identical renders dedupe. Caller is responsible for
    db.session.commit().

    If `revision` is omitted, the next available revision is assigned. Pass
    it explicitly when you've already used the same value while rendering
    (e.g., to fill a {{revision}} placeholder), so the printed and archived
    numbers match.
    """
    sha = hashlib.sha256(content).hexdigest()
    sub = _upload_root() / sha[0:2] / sha[2:4]
    sub.mkdir(parents=True, exist_ok=True)
    final = sub / sha
    if not final.exists():
        with open(final, "wb") as f:
            f.write(content)

    if revision is None:
        revision = next_po_revision(po.id)
    safe_num = (po.po_number or f"po-{po.id}").replace("/", "-").replace("\\", "-")
    filename = f"{safe_num}-rev{revision}.xlsx"

    doc = PODocument(
        po_id=po.id,
        revision=revision,
        sha256=sha,
        original_filename=filename,
        mime_type=mime_type,
        size_bytes=len(content),
        template_name=template_name,
        generated_by=generated_by,
    )
    db.session.add(doc)
    return doc


def po_document_path(doc: PODocument) -> Path:
    sha = doc.sha256
    return _upload_root() / sha[0:2] / sha[2:4] / sha


def delete_po_document(doc: PODocument) -> None:
    """Remove the DB row and the on-disk blob if no other rows reference it."""
    sha = doc.sha256
    path = po_document_path(doc)
    db.session.delete(doc)
    db.session.flush()
    still_referenced = (
        db.session.query(PODocument.id).filter_by(sha256=sha).first() is not None
        or db.session.query(Attachment.id).filter_by(sha256=sha).first() is not None
    )
    if not still_referenced and path.exists():
        try:
            path.unlink()
        except OSError:
            pass


def set_item_state(item: Item, new_state: str) -> None:
    """Manual override (used for approved / cancelled / un-cancelling)."""
    states = current_app.config["ITEM_STATES"]
    if new_state not in states:
        raise ValueError(f"Unknown state: {new_state}")
    item.state = new_state


# ---------- Attachment storage ----------

def _upload_root() -> Path:
    return Path(current_app.config["UPLOAD_DIR_RESOLVED"])


def store_attachment(stream: BinaryIO, original_filename: str,
                     mime_type: Optional[str] = None,
                     kind: str = "other",
                     uploaded_by: Optional[str] = None,
                     item_id: Optional[int] = None,
                     po_id: Optional[int] = None) -> Attachment:
    """Persist file content under uploads/<aa>/<bb>/<sha256> and return the
    Attachment row. Caller is responsible for db.session.commit()."""
    h = hashlib.sha256()
    # Buffer to disk while hashing so we don't load huge files into memory.
    tmp = io.BytesIO()
    while True:
        chunk = stream.read(1024 * 64)
        if not chunk:
            break
        h.update(chunk)
        tmp.write(chunk)
    sha = h.hexdigest()
    size = tmp.tell()

    sub = _upload_root() / sha[0:2] / sha[2:4]
    sub.mkdir(parents=True, exist_ok=True)
    final = sub / sha
    if not final.exists():
        with open(final, "wb") as f:
            tmp.seek(0)
            shutil.copyfileobj(tmp, f)

    att = Attachment(
        sha256=sha,
        original_filename=original_filename,
        mime_type=mime_type,
        size_bytes=size,
        kind=kind,
        uploaded_by=uploaded_by,
        item_id=item_id,
        po_id=po_id,
    )
    db.session.add(att)
    return att


def attachment_path(att: Attachment) -> Path:
    sha = att.sha256
    return _upload_root() / sha[0:2] / sha[2:4] / sha


def delete_attachment(att: Attachment) -> None:
    """Remove DB row and the on-disk blob if no other rows reference it."""
    sha = att.sha256
    path = attachment_path(att)
    db.session.delete(att)
    db.session.flush()
    still_referenced = (
        db.session.query(Attachment.id).filter_by(sha256=sha).first() is not None
    )
    if not still_referenced and path.exists():
        try:
            path.unlink()
        except OSError:
            pass  # leave orphaned blob; better than crashing


# ---------- Tags ----------

def get_or_create_tag(name: str) -> Tag:
    name = name.strip()
    if not name:
        raise ValueError("Empty tag")
    tag = db.session.query(Tag).filter_by(name=name).first()
    if tag is None:
        tag = Tag(name=name)
        db.session.add(tag)
        db.session.flush()
    return tag


def apply_tags(item: Item, names: Iterable[str]) -> None:
    cleaned = [n.strip() for n in names if n and n.strip()]
    item.tags = [get_or_create_tag(n) for n in cleaned]


# ---------- xlsx PO template rendering ----------

# Patterns we recognise inside a cell's text.
_PLACEHOLDER_RE = re.compile(r"\{\{\s*([a-zA-Z0-9_.]+)\s*\}\}")
_LOOP_OPEN_RE = re.compile(r"\{\{\s*#\s*items\s*\}\}")
_LOOP_CLOSE_RE = re.compile(r"\{\{\s*/\s*items\s*\}\}")
_IF_OPEN_RE = re.compile(r"\{\{\s*#\s*if\s+([a-zA-Z0-9_.]+)\s*\}\}")
_ELSE_RE = re.compile(r"\{\{\s*else\s*\}\}")
_IF_CLOSE_RE = re.compile(r"\{\{\s*/\s*if\s*\}\}")


def render_po_xlsx(po: PurchaseOrder, template_bytes: bytes,
                   revision: Optional[int] = None) -> bytes:
    """Render a PO into a copy of the supplied xlsx template.

    Two replacement modes coexist:

    1. Named cells: any cell whose text contains {{name}} is substituted
       in place. Surrounding formatting on that cell is preserved.

    2. Loop region: rows from the row containing {{#items}} through the
       row containing {{/items}} (exclusive of marker rows) are treated
       as a per-line template. They are duplicated once per PO line
       and the rows that originally followed are shifted down. The
       marker rows themselves are deleted.

    Lines are ordered by their stable line_no so the # column in the rendered
    document matches what the user sees in the web UI.

    Returns the rendered xlsx as bytes.
    """
    wb = load_workbook(io.BytesIO(template_bytes))

    context = _po_context(po, revision=revision)

    for ws in wb.worksheets:
        _render_loop_region(ws, po)
        _render_named_cells(ws, context)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _po_context(po: PurchaseOrder, revision: Optional[int] = None) -> dict:
    return {
        "po_number": po.po_number or "",
        "vendor": po.vendor or "",
        "ship_to": po.ship_to or "",
        "notes": po.notes or "",
        "date": (po.ordered_at or po.created_at).strftime("%Y-%m-%d")
                 if (po.ordered_at or po.created_at) else "",
        "total": po.total,
        "revision": revision if revision is not None else "",
    }


def _line_context(line: POLine) -> dict:
    item = line.item
    return {
        "item.name": item.name if item else "",
        "item.description": (item.description if item else "") or "",
        "item.model": (item.model if item else "") or "",
        "item.vendor": (item.vendor if item else "") or "",
        "item.vendor_sku": (item.vendor_sku if item else "") or "",
        "item.url": (item.url if item else "") or "",
        "item.qty": line.qty,
        "item.unit_cost": line.unit_cost,
        "item.line_total": line.line_total,
        # Stable per-PO line number set when the line was added (or, on legacy
        # DBs, backfilled by migration). Matches the "#" shown in the UI.
        "item.index": line.line_no or 0,
        "item.notes": line.notes or "",
    }


def _resolve_conditionals(text: str, ctx: dict) -> str:
    """Process {{#if var}}...{{else}}...{{/if}} blocks within a cell value.

    The {{else}} branch is optional; omitting it means an empty string when
    the condition is false. Blocks do not nest.
    """
    result = []
    pos = 0
    while pos < len(text):
        m = _IF_OPEN_RE.search(text, pos)
        if m is None:
            result.append(text[pos:])
            break
        result.append(text[pos:m.start()])
        var = m.group(1)
        end_m = _IF_CLOSE_RE.search(text, m.end())
        if end_m is None:
            # Malformed block — leave the rest untouched.
            result.append(text[m.start():])
            break
        inner = text[m.end():end_m.start()]
        else_m = _ELSE_RE.search(inner)
        condition = bool(ctx.get(var))
        if else_m:
            branch = inner[:else_m.start()] if condition else inner[else_m.end():]
        else:
            branch = inner if condition else ""
        result.append(branch)
        pos = end_m.end()
    return "".join(result)


def _substitute(text: str, ctx: dict) -> str:
    text = _resolve_conditionals(text, ctx)

    def repl(m):
        key = m.group(1)
        val = ctx.get(key)
        if val is None:
            return m.group(0)  # leave unknown placeholders untouched
        return str(val)
    return _PLACEHOLDER_RE.sub(repl, text)


def _cell_substitute(cell: Cell, ctx: dict) -> None:
    if not isinstance(cell.value, str):
        return
    new = _substitute(cell.value, ctx)
    if new != cell.value:
        # If the new value is a number and the placeholder was the only
        # content, coerce to a number so the spreadsheet can sum it.
        stripped = new.strip()
        if _looks_numeric(stripped):
            try:
                cell.value = float(stripped) if "." in stripped else int(stripped)
                return
            except ValueError:
                pass
        cell.value = new


def _looks_numeric(s: str) -> bool:
    return bool(re.fullmatch(r"-?\d+(\.\d+)?", s))


def _render_named_cells(ws: Worksheet, ctx: dict) -> None:
    for row in ws.iter_rows():
        for cell in row:
            _cell_substitute(cell, ctx)


def _find_loop_region(ws: Worksheet) -> Optional[Tuple[int, int]]:
    """Return (open_row, close_row) of the first {{#items}}/{{/items}} pair,
    or None if absent. Both are 1-indexed Excel row numbers."""
    open_row = None
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            if open_row is None and _LOOP_OPEN_RE.search(cell.value):
                open_row = cell.row
            elif open_row is not None and _LOOP_CLOSE_RE.search(cell.value):
                return open_row, cell.row
    return None


def _render_loop_region(ws: Worksheet, po: PurchaseOrder) -> None:
    region = _find_loop_region(ws)
    if region is None:
        return
    open_row, close_row = region
    template_rows = list(range(open_row + 1, close_row))
    template_height = len(template_rows)

    # --- Snapshot ALL merged ranges before any mutations. ---
    # openpyxl's insert_rows/delete_rows have well-known bugs that corrupt
    # merged cell ranges. We take full control: unmerge everything first,
    # then re-apply adjusted ranges ourselves after the mutations.
    all_merges = [
        (mr.min_row, mr.max_row, mr.min_col, mr.max_col)
        for mr in ws.merged_cells.ranges
    ]
    for mr in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(mr))

    # Split merges into two buckets:
    #   template_merges — fully inside the template rows; replicated per line.
    #   external_merges — outside the loop block; row-shifted by the net delta.
    tmpl_start, tmpl_end = open_row + 1, close_row - 1
    template_merges: list[tuple[int, int, int, int]] = []  # offsets from tmpl_start
    external_merges: list[tuple[int, int, int, int]] = []
    for (min_r, max_r, min_c, max_c) in all_merges:
        if min_r >= tmpl_start and max_r <= tmpl_end:
            template_merges.append((min_r - tmpl_start, max_r - tmpl_start, min_c, max_c))
        else:
            external_merges.append((min_r, max_r, min_c, max_c))

    # Capture template row data (values + styles + row height).
    max_col = ws.max_column
    template = []
    for r in template_rows:
        rd = ws.row_dimensions.get(r)
        row_snapshot = {
            "height": rd.height if rd else None,
            "cells": [],
        }
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row_snapshot["cells"].append({
                "value": cell.value,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
            })
        template.append(row_snapshot)

    lines = sorted(po.lines, key=_render_sort_key)

    rows_to_remove = close_row - open_row + 1  # inclusive of both markers
    ws.delete_rows(open_row, rows_to_remove)

    insert_count = template_height * len(lines)

    if not lines:
        _reapply_external_merges(ws, external_merges, open_row, close_row, -rows_to_remove)
        return

    if insert_count > 0:
        ws.insert_rows(open_row, amount=insert_count)

    write_row = open_row
    for line in lines:
        ctx = _line_context(line)
        line_base = write_row
        for tmpl_row in template:
            if tmpl_row["height"] is not None:
                ws.row_dimensions[write_row].height = tmpl_row["height"]
            for col_idx, src in enumerate(tmpl_row["cells"], start=1):
                cell = ws.cell(row=write_row, column=col_idx)
                value = src["value"]
                if isinstance(value, str):
                    value = _substitute(value, ctx)
                    if _looks_numeric(value.strip()):
                        try:
                            value = (float(value) if "." in value
                                     else int(value))
                        except ValueError:
                            pass
                cell.value = value
                cell.font = copy(src["font"])
                cell.fill = copy(src["fill"])
                cell.border = copy(src["border"])
                cell.alignment = copy(src["alignment"])
                cell.number_format = src["number_format"]
            write_row += 1
        for (off_min_r, off_max_r, min_c, max_c) in template_merges:
            ws.merge_cells(
                start_row=line_base + off_min_r,
                end_row=line_base + off_max_r,
                start_column=min_c,
                end_column=max_c,
            )

    delta = insert_count - rows_to_remove
    _reapply_external_merges(ws, external_merges, open_row, close_row, delta)


def _reapply_external_merges(
    ws: Worksheet,
    external_merges: list[tuple[int, int, int, int]],
    open_row: int,
    close_row: int,
    delta: int,
) -> None:
    """Re-apply merged cell ranges that were outside the loop region.

    Ranges entirely above open_row are unchanged. Ranges that started after
    close_row are shifted by delta. Ranges that overlapped the loop block
    itself are silently dropped — they were invalid template constructs.
    """
    for (min_r, max_r, min_c, max_c) in external_merges:
        if max_r < open_row:
            pass  # entirely above — no adjustment needed
        elif min_r > close_row:
            min_r += delta
            max_r += delta
        else:
            continue  # overlapped the loop block — drop it
        ws.merge_cells(start_row=min_r, end_row=max_r, start_column=min_c, end_column=max_c)


# ---------- xlsx <-> x-spreadsheet JSON (template editor) ----------
#
# The template editor in the web UI uses x-spreadsheet, a vanilla-JS
# spreadsheet component. It loads and emits a JSON document with this rough
# shape:
#
#   {
#     "name": "Sheet1",
#     "rows": { "0": { "cells": { "0": {"text": "Hi", "style": 0} },
#                       "height": 24 }, "len": 100 },
#     "cols": { "0": {"width": 120}, "len": 26 },
#     "merges": ["A1:B2"],
#     "styles": [ {"font": {"bold": true}, "bgcolor": "#fff", ...} ]
#   }
#
# We only model the subset of cell properties that round-trip cleanly: text
# (including formulas as "=…"), basic font, alignment, background fill,
# number format, merges, column widths, row heights. Anything else from the
# source xlsx is dropped on the edit/save round trip — documented in the UI.

_XSS_PX_PER_WIDTH_UNIT = 7  # rough Excel char-width → pixels conversion


def xlsx_bytes_to_xspreadsheet(xlsx_bytes: bytes) -> dict:
    """Convert an .xlsx workbook (bytes) into x-spreadsheet JSON for editing."""
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    palette = _load_theme_palette(wb)
    return _worksheet_to_xspreadsheet(ws, palette)


def _worksheet_to_xspreadsheet(ws: Worksheet, palette: dict) -> dict:
    rows: dict[str, Any] = {}
    styles: list[dict] = []
    style_cache: dict[tuple, int] = {}

    max_r = max(ws.max_row or 1, 1)
    max_c = max(ws.max_column or 1, 1)

    for r in range(1, max_r + 1):
        row_cells: dict[str, Any] = {}
        for c in range(1, max_c + 1):
            cell = ws.cell(row=r, column=c)
            val = cell.value
            has_text = val is not None and val != ""
            key, style_obj = _cell_style_signature(cell, palette)
            # Emit a cell if it has text OR a non-default style. Blank-but-
            # styled cells (coloured input boxes, bordered spacers, the BILL
            # TO block, the purple barcode cells) MUST be kept — skipping
            # them was why large swathes of the form rendered uncoloured.
            if not has_text and style_obj is None:
                continue
            cell_data: dict[str, Any] = {
                "text": _xlsx_value_to_text(val) if has_text else "",
            }
            if style_obj is not None:
                if key not in style_cache:
                    style_cache[key] = len(styles)
                    styles.append(style_obj)
                cell_data["style"] = style_cache[key]
            row_cells[str(c - 1)] = cell_data
        rd = ws.row_dimensions.get(r)
        height = rd.height if rd is not None and rd.height else None
        if row_cells or height:
            row_obj: dict[str, Any] = {}
            if row_cells:
                row_obj["cells"] = row_cells
            if height:
                row_obj["height"] = float(height)
            rows[str(r - 1)] = row_obj
    rows["len"] = max(max_r + 10, 100)

    # Emit a width for EVERY column up to max_c, not just ones with an
    # explicit width. Excel's default is ~8.43 chars; if we leave default
    # columns out, x-spreadsheet falls back to its own (different) default
    # and the whole form's proportions drift — narrow columns wrapped text
    # ("MODEL/S KU") while others stretched.
    default_w = None
    sf = getattr(ws, "sheet_format", None)
    if sf is not None and getattr(sf, "defaultColWidth", None):
        default_w = sf.defaultColWidth
    if not default_w:
        default_w = 8.43

    cols: dict[str, Any] = {}
    for c in range(1, max_c + 1):
        letter = get_column_letter(c)
        dim = ws.column_dimensions.get(letter)
        w = dim.width if dim is not None and dim.width else default_w
        cols[str(c - 1)] = {
            "width": int(round(w * _XSS_PX_PER_WIDTH_UNIT + 5))
        }
    cols["len"] = max(max_c + 5, 26)

    merges = [str(mr) for mr in ws.merged_cells.ranges]

    # x-spreadsheet needs a merge declared in TWO places: the `merges` array
    # (drives range selection) AND a `merge: [extraRows, extraCols]` property
    # on the anchor (top-left) cell (drives the grid actually spanning the
    # cells). Without the latter the range highlights on click but renders as
    # separate cells — exactly the inconsistency this annotation fixes.
    for mr in ws.merged_cells.ranges:
        min_r, min_c = mr.min_row, mr.min_col
        extra_rows = mr.max_row - min_r
        extra_cols = mr.max_col - min_c
        if extra_rows == 0 and extra_cols == 0:
            continue
        rkey, ckey = str(min_r - 1), str(min_c - 1)
        row_obj = rows.setdefault(rkey, {})
        cells = row_obj.setdefault("cells", {})
        anchor = cells.setdefault(ckey, {"text": ""})
        anchor["merge"] = [extra_rows, extra_cols]

    return {
        "name": (ws.title or "Sheet1")[:31],
        "rows": rows,
        "cols": cols,
        "merges": merges,
        "styles": styles,
    }


def _xlsx_value_to_text(val: Any) -> str:
    """Render an openpyxl cell value as a string suitable for x-spreadsheet.

    Formulas come back from openpyxl as their text starting with '=' (when the
    workbook is loaded without data_only=True), so str(val) is correct.
    """
    if isinstance(val, bool):
        return "TRUE" if val else "FALSE"
    if isinstance(val, (int, float)):
        # Integers print without trailing .0; floats keep their natural repr.
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val)
    if isinstance(val, dt.datetime):
        return val.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(val, dt.date):
        return val.strftime("%Y-%m-%d")
    return str(val)


def _cell_style_signature(
    cell: Cell, palette: dict,
) -> tuple[Optional[tuple], Optional[dict]]:
    """Return (cache_key, style_dict) for this cell, or (None, None) if default.

    The cache key is a hashable shape used to dedupe identical styles in the
    output. The style_dict matches x-spreadsheet's expected shape.
    """
    font = cell.font
    align = cell.alignment
    fill = cell.fill
    fmt = cell.number_format

    bold = bool(font and font.bold)
    italic = bool(font and font.italic)
    size = int(font.size) if font and font.size else None
    name = font.name if font and font.name else None
    fg = _color_to_hex(font.color, palette) if font else None

    halign = align.horizontal if align and align.horizontal in {
        "left", "center", "right"
    } else None
    valign = align.vertical if align and align.vertical in {
        "top", "middle", "bottom"
    } else None
    # openpyxl uses "center" while x-spreadsheet uses "middle" for vertical.
    if valign == "center":
        valign = "middle"
    wrap = bool(align and align.wrap_text)

    bg = None
    if fill is not None and fill.patternType == "solid":
        bg = (_color_to_hex(fill.fgColor, palette)
              or _color_to_hex(fill.bgColor, palette))
    elif fill is not None and fill.patternType:
        # Non-solid patterns (gray125, lightUp, etc.) — approximate with the
        # pattern's foreground so shaded cells at least show *something*.
        bg = _color_to_hex(fill.fgColor, palette)

    has_format = fmt and fmt not in {"General", "general", None}
    xss_fmt = _excel_to_xss_format(fmt) if has_format else None

    border_obj = _cell_border_to_xss(cell, palette)

    if not any([bold, italic, size, name, fg, halign, valign, wrap, bg,
                xss_fmt, has_format, border_obj]):
        return None, None

    style: dict[str, Any] = {}
    font_obj: dict[str, Any] = {}
    if bold:
        font_obj["bold"] = True
    if italic:
        font_obj["italic"] = True
    if size:
        font_obj["size"] = size
    if name:
        font_obj["name"] = name
    if font_obj:
        style["font"] = font_obj
    if fg:
        style["color"] = fg
    if bg:
        style["bgcolor"] = bg
    if halign:
        style["align"] = halign
    if valign:
        style["valign"] = valign
    if wrap:
        style["textwrap"] = True
    if xss_fmt:
        # Must be one of x-spreadsheet's known format keys, otherwise the
        # editor crashes during render (it does formats[style.format].render).
        style["format"] = xss_fmt
    if has_format:
        # Stash the original Excel format string under a non-conflicting key
        # so we can restore it exactly on save (x-spreadsheet only looks up
        # `style.format`, so unknown keys are safe).
        style["xlsxFormat"] = fmt
    if border_obj:
        style["border"] = border_obj

    key = (
        bold, italic, size, name, fg, halign, valign, wrap, bg,
        fmt if has_format else None,
        _border_cache_key(border_obj),
    )
    return key, style


# Border styles x-spreadsheet renders natively. Anything else from openpyxl
# is mapped to the closest visual equivalent.
_XSS_BORDER_STYLES = {"thin", "medium", "thick", "dashed", "dotted", "double"}

# openpyxl side style → x-spreadsheet side style.
_OPENPYXL_TO_XSS_BORDER = {
    "thin": "thin", "medium": "medium", "thick": "thick",
    "dashed": "dashed", "dotted": "dotted", "double": "double",
    "hair": "thin", "dashDot": "dashed", "dashDotDot": "dashed",
    "mediumDashed": "medium", "mediumDashDot": "medium",
    "mediumDashDotDot": "medium", "slantDashDot": "dashed",
}


def _side_to_xss(side: Any, palette: dict) -> Optional[list]:
    """Convert an openpyxl Side to x-spreadsheet's ['style', '#color'] tuple."""
    if side is None or not getattr(side, "style", None):
        return None
    style = _OPENPYXL_TO_XSS_BORDER.get(side.style, "thin")
    color = _color_to_hex(side.color, palette) or "#000000"
    return [style, color]


def _cell_border_to_xss(cell: Cell, palette: dict) -> Optional[dict]:
    """Return an x-spreadsheet border dict, or None if the cell has no borders."""
    border = cell.border
    if border is None:
        return None
    top = _side_to_xss(border.top, palette)
    bottom = _side_to_xss(border.bottom, palette)
    left = _side_to_xss(border.left, palette)
    right = _side_to_xss(border.right, palette)
    if not any([top, bottom, left, right]):
        return None
    out: dict[str, list] = {}
    if top:    out["top"] = top
    if bottom: out["bottom"] = bottom
    if left:   out["left"] = left
    if right:  out["right"] = right
    return out


def _border_cache_key(border_obj: Optional[dict]) -> Optional[tuple]:
    if not border_obj:
        return None
    return tuple(
        (k, tuple(v)) for k, v in sorted(border_obj.items())
    )


# Mapping between Excel number formats and x-spreadsheet's named formats.
# x-spreadsheet's known names: normal, text, number, percent, rmb, usd, eur,
# date, time, datetime, duration. Any other value passed as style.format
# causes a runtime crash, so we MUST translate.
_XSS_NAMED_FORMATS = {
    "normal", "text", "number", "percent",
    "rmb", "usd", "eur", "date", "time", "datetime", "duration",
}


def _excel_to_xss_format(fmt: Optional[str]) -> Optional[str]:
    """Map an Excel number format string to x-spreadsheet's nearest named format.

    Returns None for formats we don't recognise — the caller should then omit
    `style.format` entirely (the raw Excel format string is preserved under
    `style.xlsxFormat` so the round-trip back to xlsx is lossless).
    """
    if not fmt:
        return None
    f = fmt.lower()
    if f in _XSS_NAMED_FORMATS:
        return f
    if "$" in fmt:
        return "usd"
    if "€" in fmt:
        return "eur"
    if "¥" in fmt or "rmb" in f:
        return "rmb"
    if "%" in fmt:
        return "percent"
    # Date/time detection: openpyxl format strings use y/m/d/h tokens.
    has_date = any(tok in f for tok in ("yyyy", "yy", "mmm", "dd", "m/d"))
    has_time = "h" in f and "m" in f  # "hh:mm" patterns
    if has_date and has_time:
        return "datetime"
    if has_date:
        return "date"
    if has_time:
        return "time"
    if any(ch in fmt for ch in "0#"):
        return "number"
    return None


# Reverse mapping used when saving the editor's data back to xlsx.
_XSS_TO_EXCEL_FORMAT = {
    "number":   "#,##0.00",
    "percent":  "0.00%",
    "usd":      '"$"#,##0.00',
    "eur":      '"€"#,##0.00',
    "rmb":      '"¥"#,##0.00',
    "date":     "yyyy-mm-dd",
    "time":     "hh:mm:ss",
    "datetime": "yyyy-mm-dd hh:mm:ss",
}


def _argb_to_hex(rgb: Any) -> Optional[str]:
    """openpyxl stores colors as 'AARRGGBB'. Return '#RRGGBB' or None."""
    if not isinstance(rgb, str):
        return None
    s = rgb.strip()
    if len(s) == 8:
        return "#" + s[2:].lower()
    if len(s) == 6:
        return "#" + s.lower()
    return None


# Default Office theme palette (indices match openpyxl's Color(theme=N)).
# Modern xlsx files store fills/fonts as theme refs rather than raw RGB; if
# we can't extract a real RGB from the cell, fall back to this table so the
# editor at least shows roughly the right colour. (Tint is intentionally
# ignored — preserving exact shades would require parsing the theme XML.)
_OFFICE_THEME_RGB = {
    0: "#ffffff", 1: "#000000",
    2: "#e7e6e6", 3: "#44546a",
    4: "#5b9bd5", 5: "#ed7d31",
    6: "#a5a5a5", 7: "#ffc000",
    8: "#4472c4", 9: "#70ad47",
    10: "#0563c1", 11: "#954f72",
}

# Standard Excel indexed colour table (subset — entries 0-63 cover everything
# real-world templates use; values come from the OOXML spec).
_INDEXED_COLOR_RGB = {
    0:  "#000000", 1:  "#ffffff", 2:  "#ff0000", 3:  "#00ff00",
    4:  "#0000ff", 5:  "#ffff00", 6:  "#ff00ff", 7:  "#00ffff",
    8:  "#000000", 9:  "#ffffff", 10: "#ff0000", 11: "#00ff00",
    12: "#0000ff", 13: "#ffff00", 14: "#ff00ff", 15: "#00ffff",
    16: "#800000", 17: "#008000", 18: "#000080", 19: "#808000",
    20: "#800080", 21: "#008080", 22: "#c0c0c0", 23: "#808080",
    24: "#9999ff", 25: "#993366", 26: "#ffffcc", 27: "#ccffff",
    28: "#660066", 29: "#ff8080", 30: "#0066cc", 31: "#ccccff",
    32: "#000080", 33: "#ff00ff", 34: "#ffff00", 35: "#00ffff",
    36: "#800080", 37: "#800000", 38: "#008080", 39: "#0000ff",
    40: "#00ccff", 41: "#ccffff", 42: "#ccffcc", 43: "#ffff99",
    44: "#99ccff", 45: "#ff99cc", 46: "#cc99ff", 47: "#ffcc99",
    48: "#3366ff", 49: "#33cccc", 50: "#99cc00", 51: "#ffcc00",
    52: "#ff9900", 53: "#ff6600", 54: "#666699", 55: "#969696",
    56: "#003366", 57: "#339966", 58: "#003300", 59: "#333300",
    60: "#993300", 61: "#993366", 62: "#333399", 63: "#333333",
    64: "#000000", 65: "#ffffff",
}


# The clrScheme children appear in the theme XML as dk1, lt1, dk2, lt2,
# accent1-6, hlink, folHlink. openpyxl's Color(theme=N) index, however,
# swaps the first two pairs: 0=lt1, 1=dk1, 2=lt2, 3=dk2, then 4..9 accents,
# 10 hlink, 11 folHlink. This maps XML position -> theme index.
_THEME_XML_ORDER_TO_INDEX = {0: 1, 1: 0, 2: 3, 3: 2,
                             4: 4, 5: 5, 6: 6, 7: 7, 8: 8, 9: 9,
                             10: 10, 11: 11}

_A_NS = "{http://schemas.openxmlformats.org/drawingml/2006/main}"


def _load_theme_palette(wb: Any) -> dict:
    """Parse the workbook's theme1.xml into {theme_index: '#rrggbb'}.

    Falls back to the default Office palette for any slot we can't read.
    Custom corporate templates frequently ship their own theme, so reading
    the real one (rather than assuming Office defaults) is what makes their
    fills/borders show the right colour.
    """
    palette = dict(_OFFICE_THEME_RGB)
    raw = getattr(wb, "loaded_theme", None)
    if not raw:
        return palette
    try:
        if isinstance(raw, bytes):
            root = ET.fromstring(raw)
        else:
            root = ET.fromstring(raw.encode("utf-8"))
    except ET.ParseError:
        return palette

    scheme = root.find(f".//{_A_NS}clrScheme")
    if scheme is None:
        return palette

    for pos, child in enumerate(list(scheme)):
        idx = _THEME_XML_ORDER_TO_INDEX.get(pos)
        if idx is None:
            continue
        srgb = child.find(f"{_A_NS}srgbClr")
        sysclr = child.find(f"{_A_NS}sysClr")
        hexval = None
        if srgb is not None and srgb.get("val"):
            hexval = srgb.get("val")
        elif sysclr is not None:
            hexval = sysclr.get("lastClr") or sysclr.get("val")
        if hexval and re.fullmatch(r"[0-9A-Fa-f]{6}", hexval):
            palette[idx] = "#" + hexval.lower()
    return palette


def _apply_tint(hex_color: str, tint: float) -> str:
    """Apply the OOXML tint factor to a '#rrggbb' colour.

    Excel stores subtle shades ("white, darker 15%") as a base theme colour
    plus a tint in [-1, 1]. Ignoring the tint is why a light-grey fill was
    coming through as pure white (invisible). Algorithm per ECMA-376: adjust
    the HSL luminance.
    """
    if not tint:
        return hex_color
    s = hex_color.lstrip("#")
    if len(s) != 6:
        return hex_color
    r, g, b = (int(s[i:i + 2], 16) / 255.0 for i in (0, 2, 4))
    h, l, sat = colorsys.rgb_to_hls(r, g, b)
    if tint < 0:
        l = l * (1.0 + tint)
    else:
        l = l * (1.0 - tint) + tint
    l = min(1.0, max(0.0, l))
    r, g, b = colorsys.hls_to_rgb(h, l, sat)
    return "#%02x%02x%02x" % (
        round(r * 255), round(g * 255), round(b * 255),
    )


def _color_to_hex(color: Any, palette: dict) -> Optional[str]:
    """Resolve an openpyxl Color to a '#rrggbb' string, regardless of its type.

    Handles type='rgb', 'theme' (resolved against the workbook's actual
    palette, with tint applied), and 'indexed' (legacy palette). Returns
    None if we can't recover any usable colour.
    """
    if color is None:
        return None
    ctype = getattr(color, "type", None)
    tint = getattr(color, "tint", 0.0) or 0.0

    if ctype == "rgb":
        base = _argb_to_hex(getattr(color, "rgb", None))
        return _apply_tint(base, tint) if base else None
    if ctype == "theme":
        theme_idx = getattr(color, "theme", None)
        if isinstance(theme_idx, int):
            base = palette.get(theme_idx) or _OFFICE_THEME_RGB.get(theme_idx)
            return _apply_tint(base, tint) if base else None
        return None
    if ctype == "indexed":
        idx = getattr(color, "indexed", None)
        if isinstance(idx, int):
            base = _INDEXED_COLOR_RGB.get(idx)
            return _apply_tint(base, tint) if base else None
        return None
    # auto / unset / unknown
    base = _argb_to_hex(getattr(color, "rgb", None))
    return _apply_tint(base, tint) if base else None


def xspreadsheet_to_xlsx_bytes(data: dict) -> bytes:
    """Convert x-spreadsheet JSON back into an .xlsx workbook (bytes).

    Accepts either a single sheet dict (legacy) or a list of sheets — uses
    only the first sheet in either case, since we only support one-sheet
    templates today.
    """
    if isinstance(data, list):
        data = data[0] if data else {}
    if not isinstance(data, dict):
        data = {}

    wb = Workbook()
    ws = wb.active
    ws.title = (str(data.get("name") or "Sheet1"))[:31]

    styles = data.get("styles") or []
    rows = data.get("rows") or {}
    cols = data.get("cols") or {}
    merges = data.get("merges") or []

    for rk, rv in rows.items():
        if rk == "len" or not isinstance(rv, dict):
            continue
        try:
            r = int(rk) + 1
        except (TypeError, ValueError):
            continue
        cells = rv.get("cells") or {}
        for ck, cv in cells.items():
            if not isinstance(cv, dict):
                continue
            try:
                c = int(ck) + 1
            except (TypeError, ValueError):
                continue
            text = cv.get("text", "")
            style_idx = cv.get("style")
            has_style = isinstance(style_idx, int) and 0 <= style_idx < len(styles)
            # Keep blank-but-styled cells: a cell with no text but a fill or
            # border still needs to be written, otherwise coloured input
            # boxes and bordered spacers vanish on save.
            if (text == "" or text is None) and not has_style:
                continue
            cell = ws.cell(row=r, column=c)
            if text != "" and text is not None:
                cell.value = _text_to_xlsx_value(text)
            if has_style:
                _apply_style_to_cell(cell, styles[style_idx])
        height = rv.get("height")
        if height:
            try:
                ws.row_dimensions[r].height = float(height)
            except (TypeError, ValueError):
                pass

    for ck, cv in cols.items():
        if ck == "len" or not isinstance(cv, dict):
            continue
        try:
            c = int(ck) + 1
        except (TypeError, ValueError):
            continue
        width_px = cv.get("width")
        if not width_px:
            continue
        excel_width = max(2.0, (float(width_px) - 5) / _XSS_PX_PER_WIDTH_UNIT)
        ws.column_dimensions[get_column_letter(c)].width = excel_width

    for mr in merges:
        if not isinstance(mr, str):
            continue
        try:
            ws.merge_cells(mr)
        except (ValueError, TypeError):
            pass  # malformed range — skip rather than fail the save

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _text_to_xlsx_value(text: str) -> Any:
    """Coerce x-spreadsheet cell text into the appropriate openpyxl value.

    - Strings starting with '=' are preserved verbatim so openpyxl writes them
      as formulas.
    - Strings that look like plain integers/floats are coerced so the sheet
      can do arithmetic on them (otherwise SUM() over them won't add up).
    - Everything else stays as a string. Templates use {{placeholders}} which
      must remain string literals.
    """
    if not isinstance(text, str):
        return text
    if text.startswith("="):
        return text
    stripped = text.strip()
    if _looks_numeric(stripped):
        try:
            return float(stripped) if "." in stripped else int(stripped)
        except ValueError:
            return text
    return text


def _apply_style_to_cell(cell: Cell, style: dict) -> None:
    if not isinstance(style, dict):
        return
    font_src = style.get("font") or {}
    color = style.get("color")
    if font_src or color:
        cell.font = Font(
            name=font_src.get("name") or cell.font.name,
            size=font_src.get("size") or cell.font.size,
            bold=bool(font_src.get("bold")),
            italic=bool(font_src.get("italic")),
            color=_hex_to_argb(color) if color else cell.font.color,
        )

    halign = style.get("align")
    valign = style.get("valign")
    if valign == "middle":
        valign = "center"  # openpyxl spelling
    wrap = style.get("textwrap")
    if halign or valign or wrap:
        cell.alignment = Alignment(
            horizontal=halign if halign in {"left", "center", "right"} else None,
            vertical=valign if valign in {"top", "center", "bottom"} else None,
            wrap_text=bool(wrap),
        )

    bgcolor = style.get("bgcolor")
    if bgcolor:
        argb = _hex_to_argb(bgcolor)
        if argb:
            cell.fill = PatternFill("solid", fgColor=argb)

    # Number format. We stash the original Excel format string in
    # `xlsxFormat` on load so an *unchanged* format round-trips exactly.
    # But x-spreadsheet deep-clones the whole style (xlsxFormat included)
    # when the user changes the format dropdown, only overwriting
    # `style.format`. So if `format` no longer agrees with `xlsxFormat`,
    # the user changed it and we must honour `format`, not the stale stash.
    xlsx_fmt = style.get("xlsxFormat")
    xss_fmt = style.get("format")
    xss_fmt_l = xss_fmt.lower() if isinstance(xss_fmt, str) else None
    has_xlsx_fmt = isinstance(xlsx_fmt, str) and bool(xlsx_fmt)

    if xss_fmt_l in ("normal", "text", "general"):
        # User explicitly cleared the format back to plain.
        cell.number_format = "General"
    elif has_xlsx_fmt and (
        xss_fmt_l is None or _excel_to_xss_format(xlsx_fmt) == xss_fmt_l
    ):
        # Unchanged from the source workbook — restore the exact Excel format.
        cell.number_format = xlsx_fmt
    elif xss_fmt_l:
        # User picked a different named format in the editor.
        mapped = _XSS_TO_EXCEL_FORMAT.get(xss_fmt_l)
        cell.number_format = mapped or "General"

    border = style.get("border")
    if isinstance(border, dict):
        cell.border = _xss_border_to_openpyxl(border)


def _xss_side(spec: Any) -> Optional[Side]:
    """Decode ['thin', '#000000'] (x-spreadsheet shape) into an openpyxl Side."""
    if not isinstance(spec, (list, tuple)) or not spec:
        return None
    style = spec[0] if len(spec) > 0 else "thin"
    color_hex = spec[1] if len(spec) > 1 else "#000000"
    if style not in _XSS_BORDER_STYLES:
        style = "thin"
    argb = _hex_to_argb(color_hex) or "FF000000"
    return Side(style=style, color=argb)


def _xss_border_to_openpyxl(border: dict) -> Border:
    return Border(
        top=_xss_side(border.get("top")),
        bottom=_xss_side(border.get("bottom")),
        left=_xss_side(border.get("left")),
        right=_xss_side(border.get("right")),
    )


def _hex_to_argb(hex_str: Optional[str]) -> Optional[str]:
    if not isinstance(hex_str, str):
        return None
    s = hex_str.strip().lstrip("#")
    if len(s) == 6:
        return "FF" + s.upper()
    if len(s) == 8:
        return s.upper()
    return None


def sample_po_for_preview() -> SimpleNamespace:
    """Return an in-memory PurchaseOrder-shaped object for template preview.

    Duck-typed (SimpleNamespace, not a real ORM row) so it never touches the
    database. render_po_xlsx only reads attributes; it doesn't care that this
    isn't a real model instance.
    """
    items = [
        SimpleNamespace(
            name="Widget Alpha", description="Standard widget",
            model="WA-100", vendor="Acme Corp",
            vendor_sku="SKU-100", url="https://example.com/wa",
        ),
        SimpleNamespace(
            name="Gadget Beta", description="Premium gadget",
            model="GB-200", vendor="Acme Corp",
            vendor_sku="SKU-200", url="https://example.com/gb",
        ),
        SimpleNamespace(
            name="Sprocket Gamma", description="Heavy-duty sprocket",
            model="SG-300", vendor="Acme Corp",
            vendor_sku="SKU-300", url="https://example.com/sg",
        ),
    ]
    lines: list[SimpleNamespace] = []
    for i, (item, qty, cost) in enumerate(
        zip(items, [10, 2, 4], [12.50, 199.99, 45.00]), start=1
    ):
        line = SimpleNamespace(
            id=i, qty=qty, unit_cost=cost, line_no=i, notes="", item=item,
        )
        line.line_total = qty * cost
        lines.append(line)
    po = SimpleNamespace(
        po_number="PREVIEW-001",
        vendor="Acme Corp",
        ship_to="123 Main St, Anytown USA",
        notes="This is a preview rendered with sample data.",
        ordered_at=None,
        created_at=dt.datetime.now(),
        lines=lines,
    )
    po.total = sum(l.line_total for l in lines)
    return po
