"""
Generate sample_template/po_template.xlsx - a demonstration template that
shows both named placeholders and a {{#items}}/{{/items}} loop region.

Run from project root:
    python3 sample_template/build_sample.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Purchase Order"

    bold = Font(bold=True, size=12)
    title_font = Font(bold=True, size=18)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    thin = Side(style="thin", color="999999")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)

    # Title row
    ws["A1"] = "PURCHASE ORDER"
    ws["A1"].font = title_font
    ws.merge_cells("A1:F1")

    # Header block
    ws["A3"] = "PO Number:"
    ws["A3"].font = bold
    ws["B3"] = "{{po_number}}"

    ws["A4"] = "Date:"
    ws["A4"].font = bold
    ws["B4"] = "{{date}}"

    ws["A5"] = "Vendor:"
    ws["A5"].font = bold
    ws["B5"] = "{{vendor}}"

    ws["D3"] = "Ship To:"
    ws["D3"].font = bold
    ws["E3"] = "{{ship_to}}"
    ws.merge_cells("E3:F3")

    # Line-item header
    headers = ["#", "Name", "Model / SKU", "Qty", "Unit Cost", "Line Total"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=7, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.border = border
        c.alignment = Alignment(horizontal="center")

    # Loop region: row 8 marks open, row 10 marks close, row 9 is the template line.
    ws.cell(row=8, column=1, value="{{#items}}")
    ws.cell(row=9, column=1, value="{{item.index}}").border = border
    ws.cell(row=9, column=2, value="{{item.name}}").border = border
    ws.cell(row=9, column=3, value="{{item.model}}").border = border
    ws.cell(row=9, column=4, value="{{item.qty}}").border = border
    ws.cell(row=9, column=4).alignment = Alignment(horizontal="right")
    c = ws.cell(row=9, column=5, value="{{item.unit_cost}}")
    c.border = border
    c.number_format = "$#,##0.00"
    c = ws.cell(row=9, column=6, value="{{item.line_total}}")
    c.border = border
    c.number_format = "$#,##0.00"
    ws.cell(row=10, column=1, value="{{/items}}")

    # Total row
    ws.cell(row=12, column=5, value="Total:").font = bold
    c = ws.cell(row=12, column=6, value="{{total}}")
    c.font = bold
    c.number_format = "$#,##0.00"

    # Notes
    ws["A14"] = "Notes:"
    ws["A14"].font = bold
    ws["B14"] = "{{notes}}"
    ws.merge_cells("B14:F16")
    ws["B14"].alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    widths = [6, 40, 18, 8, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    out = Path(__file__).resolve().parent / "po_template.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    build()
